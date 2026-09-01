#!/usr/bin/env python3
"""
HYBRID ROCKET ACOUSTIC ANALYSIS — V5
====================================

Made for two simultaneous 2-channel acoustic datasets.

Current experiment mapping
--------------------------
Test 1, Channel 1 = 22.8 m
Test 1, Channel 2 = 23.8 m
Test 2, Channel 1 = 68.5 m
Test 2, Channel 2 = 69.5 m

Acquisition settings
--------------------
Sampling rate = 250,000 Hz
Samples analysed per channel = up to 1,200,000
Nominal record duration at 1.2M samples = 4.8 s

The script automatically:
* reads the two acoustic files used for this experiment;
* supports .xlsx/.xlsm and .csv files;
* detects Time / Channel 1 / Channel 2 columns BY NAME (fixing the previous
  column-order bug that could accidentally analyse Time as Channel 1);
* validates each acoustic channel so a time/ramp column cannot silently pass;
* automatically detects and removes hard flatline / zero-filled acquisition tails;
* uses one ignition-aligned COMMON comparison window for distance/spectral comparisons;
* detrends and band-pass filters the acoustic waveform to 20 Hz–20 kHz;
* conservatively removes isolated one-sample electrical glitches;
* limits processing to 1,200,000 samples per channel;
* reads Daq2.xlsx chamber-pressure data;
* automatically detects the static-fire pressure onset even though the DAQ and
  WaveForms timestamps are unrelated;
* event-aligns each microphone SPL history to the chamber-pressure history;
* writes all plots, alignment tables, QC diagnostics, and result tables to
  "acoustic_results";
* creates a QC-selected 1 m source-equivalent SPL estimate;
* creates a single-sided FFT amplitude spectrum for every sensor.

IMPORTANT CALIBRATION NOTE
--------------------------
The nominal GRAS 40PH sensitivity is entered as 50 mV/Pa and conditioner gain is
set to unity (0 dB). If the signal conditioner applied voltage gain, absolute
pressure/SPL must use that actual gain. Relative attenuation/spectral comparisons
are unaffected by one common calibration scale.

Required packages:
    pip install numpy pandas scipy matplotlib openpyxl

Run:
    python hybrid_acoustics_v5_qc_source_fft.py
"""

from __future__ import annotations

import csv
import math
import re
import sys
import warnings
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy import signal

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ============================================================================
# YOUR EXPERIMENT SETTINGS — ALREADY FILLED IN
# ============================================================================

SAMPLE_RATE_HZ = 250_000.0
MAX_SAMPLES = 1_200_000                    # = 4.8 s at 250 kHz

# Exact acoustic filenames visible in your Spyder / Downloads folder.
# If you rename them, change only these two lines.
TEST_1_FILE = "serp static - Copy.csv"
TEST_2_FILE = "serp static 69m.csv"

# Combustion-chamber DAQ workbook supplied with the static-fire data.
CHAMBER_PRESSURE_FILE = "Daq2.xlsx"
CHAMBER_PRESSURE_SHEET = "Untitled"
CHAMBER_PRESSURE_COLUMN = "Chamber Pressure"
CHAMBER_TIME_COLUMN = "time"

# The workbook does not specify a unit for the chamber-pressure channel.
# The script therefore does NOT guess bar/psi; plots say "DAQ units".
CHAMBER_PRESSURE_UNIT = "DAQ units"

# Channel-distance mapping supplied for your experiment.
DISTANCES_M = {
    "T1_CH1": 22.8,
    "T1_CH2": 23.8,
    "T2_CH1": 68.5,
    "T2_CH2": 69.5,
}

DISPLAY_NAMES = {
    "T1_CH1": "Test 1 CH1 — 22.8 m",
    "T1_CH2": "Test 1 CH2 — 23.8 m",
    "T2_CH1": "Test 2 CH1 — 68.5 m",
    "T2_CH2": "Test 2 CH2 — 69.5 m",
}

NEAR_SENSORS = ["T1_CH1", "T1_CH2"]
FAR_SENSORS = ["T2_CH1", "T2_CH2"]

PAIR_DEFINITIONS = [
    ("Near 1 m pair", "T1_CH1", "T1_CH2"),
    ("Far 1 m pair", "T2_CH1", "T2_CH2"),
]

# Per-channel acoustic calibration.
# These are presently the nominal GRAS 40PH values used in the previous script.
# If a channel used a different microphone or conditioner gain, change ONLY that
# channel here before treating absolute dB values as calibrated SPL.
CHANNEL_CALIBRATION = {
    "T1_CH1": {"sensitivity_mV_per_Pa": 50.0, "conditioner_gain": 1.0},
    "T1_CH2": {"sensitivity_mV_per_Pa": 50.0, "conditioner_gain": 1.0},
    "T2_CH1": {"sensitivity_mV_per_Pa": 50.0, "conditioner_gain": 1.0},
    "T2_CH2": {"sensitivity_mV_per_Pa": 50.0, "conditioner_gain": 1.0},
}
REMOVE_DC_MEAN = True

# Acoustic filtering / cleanup.
# This is the analysis band used in the paper-style plots and prevents DC drift
# or slow voltage ramps from contaminating RMS SPL.
APPLY_BANDPASS_FILTER = True
FILTER_LOW_HZ = 20.0
FILTER_HIGH_HZ = 20_000.0
FILTER_ORDER = 4
REMOVE_LINEAR_TREND = True

# Very conservative single-sample deglitching.
# It removes only an isolated point where two enormous opposite-sign jumps occur
# back-to-back. It will NOT smooth sustained rocket transients.
DESPIKE_ENABLED = True
DESPIKE_THRESHOLD_MAD = 20.0

# Hard acquisition-dropout detection. This is deliberately conservative: it only
# trims a long trailing region that is essentially constant compared with the
# active signal. Normal post-burn microphone noise is retained.
TRIM_HARD_FLATLINE_TAIL = True
FLATLINE_BLOCK_S = 0.020
FLATLINE_MIN_TAIL_S = 0.050
FLATLINE_RELATIVE_STD = 1e-6
MIN_USABLE_ACOUSTIC_DURATION_S = 0.25

# All distance / attenuation / PSD / 1/3-octave comparisons use exactly the same
# event-relative duration. This prevents a 1.1 s near recording being averaged
# against 4.8 s of a far recording.
COMMON_WINDOW_START_AFTER_IGNITION_S = 0.020
COMMON_WINDOW_END_MARGIN_S = 0.020

# Do not let a zero-filled / mathematically zero RMS window display as -6000 dB.
# Such windows are invalid acquisition data and are plotted as gaps (NaN).
SPL_ZERO_FLOOR_PA = 1e-12

# Peak graph uses a robust percentile so isolated electrical spikes do not define
# the reported rocket peak. The absolute maximum remains in the CSV table.
ROBUST_PEAK_PERCENTILE = 99.99

# Pair-delay QC: in addition to the global correlation peak, inspect a narrow
# window around the physically expected 1 m propagation delay.
EXPECTED_DELAY_SEARCH_HALF_WIDTH_S = 0.0015
PAIR_COHERENCE_BAND_HZ = (80.0, 1000.0)
PAIR_COHERENCE_WARN_BELOW = 0.50

# Analysis bandwidth.
FREQ_MIN_HZ = 20.0
FREQ_MAX_HZ = 20_000.0

# Time-history SPL uses a short moving RMS window.
SPL_WINDOW_S = 0.050                       # 50 ms
SPL_STEP_S = 0.010                         # 10 ms

# Welch PSD settings.
WELCH_NPERSEG = 65_536
WELCH_OVERLAP = 0.50

# Spectrogram settings.
SPECTROGRAM_NPERSEG = 4_096
SPECTROGRAM_OVERLAP = 0.75
SPECTROGRAM_DISPLAY_RANGE_DB = 85.0       # visual dynamic range only

# Cross-correlation/coherence settings.
MAX_CORRELATION_LAG_S = 0.010             # ±10 ms is enough for a 1 m pair
COHERENCE_NPERSEG = 65_536

# Paper-style processing.
SAVGOL_WINDOW_POINTS = 31
SAVGOL_POLYORDER = 3
REFLECTION_CORRECTION_DB = (3.0, 6.0)

# Distance-correct spectra to this reference distance.
DISTANCE_REFERENCE_M = 1.0

# Source-equivalent estimate.  A point-source model is singular at r=0, so the
# script reports an equivalent FREE-FIELD SPL at 1 m from the acoustic source.
# This is a rough back-propagation only: reflections, vertical firing geometry,
# ground interaction, directivity and non-point-source plume effects are ignored.
SOURCE_EQUIVALENT_REFERENCE_M = 1.0
SOURCE_ESTIMATE_SPL_WINDOW_S = 0.010
SOURCE_ESTIMATE_SPL_STEP_S = 0.005
SOURCE_ESTIMATE_REQUIRE_FULL_RECORD = True
SOURCE_ESTIMATE_MIN_COHERENCE = 0.50
SOURCE_ESTIMATE_MAX_DELAY_ERROR_MS = 1.5
SOURCE_ESTIMATE_SHOW_EXCLUDED_CHANNELS = True

# Direct FFT output requested in addition to Welch PSD. For broadband rocket noise,
# PSD is normally the better comparison metric because a raw FFT amplitude depends
# on record length / bin width. The FFT plot is nevertheless useful for spotting
# discrete tones and resonances.
FFT_MIN_HZ = 20.0
FFT_MAX_HZ = 20_000.0
FFT_USE_COMMON_IGNITION_WINDOW = True
FFT_EXPORT_LINEAR_PA = True

# Pair-health QC. Two microphones only 1 m apart should not normally disagree by
# huge broadband RMS levels during the same event. Large mismatch is flagged rather
# than being 'filtered away'.
PAIR_LEVEL_MISMATCH_WARN_DB = 4.0
PAIR_DELAY_ERROR_WARN_MS = 1.5

# Air properties.
SPEED_OF_SOUND_M_S = 343.0

# Pressure/acoustic event alignment.
# Absolute clocks are not shared, so the script detects the same ignition event in
# both recordings and converts both to event-relative time.
PRESSURE_SMOOTH_WINDOW_S = 0.10
PRESSURE_ONSET_FRACTION_OF_EVENT = 0.08
PRESSURE_ONSET_SUSTAIN_S = 0.05

ACOUSTIC_ONSET_ENVELOPE_WINDOW_S = 0.010
ACOUSTIC_ONSET_FRACTION_OF_EVENT = 0.12
ACOUSTIC_ONSET_SUSTAIN_S = 0.010
ALIGNMENT_SPL_WINDOW_S = 0.010            # 10 ms, matches 100 Hz DAQ timing scale
ALIGNMENT_SPL_STEP_S = 0.005

# Output folder.
OUTPUT_FOLDER_NAME = "acoustic_results_run"

# Every execution creates a NEW sequential folder:
#   acoustic_results_run_001
#   acoustic_results_run_002
#   acoustic_results_run_003
# etc.
# Existing result folders are never overwritten.


# ============================================================================
# CONSTANTS
# ============================================================================

P_REF = 20e-6
EPS = np.finfo(float).tiny


# ============================================================================
# GENERAL HELPERS
# ============================================================================

def safe_name(s: str) -> str:
    return "".join(c if c.isalnum() or c in "-_." else "_" for c in s)


def db10(x):
    return 10.0 * np.log10(np.maximum(np.asarray(x), EPS))


def pressure_to_spl(p_rms):
    return 20.0 * np.log10(np.maximum(np.asarray(p_rms), EPS) / P_REF)


def normalise_header(value) -> str:
    s = "" if value is None else str(value)
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def is_time_header(s: str) -> bool:
    s = normalise_header(s)
    return (
        "time" in s
        or s in {"t", "seconds", "second", "sec", "s"}
    )


def channel_number_from_header(s: str) -> Optional[int]:
    s = normalise_header(s)
    # Common WaveForms / oscilloscope variants:
    patterns = {
        1: [
            r"\bchannel\s*1\b", r"\bch\s*1\b", r"\bc1\b",
            r"\bscope\s*1\b", r"\b1\s*\(v\)"
        ],
        2: [
            r"\bchannel\s*2\b", r"\bch\s*2\b", r"\bc2\b",
            r"\bscope\s*2\b", r"\b2\s*\(v\)"
        ],
    }
    for n, pats in patterns.items():
        if any(re.search(p, s) for p in pats):
            return n
    return None


def voltage_scale_from_header(header: str) -> float:
    """
    Convert numerical column units to volts.
    If column says mV -> 0.001. If V or unknown -> 1.
    """
    s = normalise_header(header)
    if "mv" in s or "millivolt" in s:
        return 1e-3
    if "uv" in s or "µv" in s or "microvolt" in s:
        return 1e-6
    return 1.0


def calibrate_to_pa(
    raw: np.ndarray,
    sensor_name: str,
    voltage_scale: float = 1.0
) -> np.ndarray:
    cal = CHANNEL_CALIBRATION[sensor_name]
    sensitivity_v_per_pa = float(cal["sensitivity_mV_per_Pa"]) / 1000.0
    gain = float(cal.get("conditioner_gain", 1.0))
    p = np.asarray(raw, dtype=np.float64) * voltage_scale
    p = p / (sensitivity_v_per_pa * gain)
    if REMOVE_DC_MEAN and len(p):
        p = p - np.nanmean(p)
    return p


def validate_raw_acoustic_signal(raw: np.ndarray, label: str) -> Dict[str, float]:
    """
    Detect the exact failure that produced the previous plots: a Time column was
    accidentally treated as Channel 1. Also reports clipping/flatline indicators.
    """
    x = np.asarray(raw, dtype=float)
    x = x[np.isfinite(x)]
    if len(x) < 100:
        raise ValueError(f"{label}: too few finite samples ({len(x)}).")

    # Correlation with sample index is a strong time/ramp detector.
    stride = max(1, len(x)//10000)
    xs = x[::stride]
    idx = np.arange(len(xs), dtype=float)
    if np.std(xs) > 0:
        index_corr = float(np.corrcoef(idx, xs)[0, 1])
    else:
        index_corr = 0.0

    dx = np.diff(xs)
    if len(dx):
        monotonic_fraction = float(
            max(np.mean(dx >= 0), np.mean(dx <= 0))
        )
    else:
        monotonic_fraction = 1.0

    xmin, xmax = float(np.min(x)), float(np.max(x))
    span = xmax - xmin
    atol = max(abs(span)*1e-12, 1e-15)
    clip_fraction = float(
        np.mean(np.isclose(x, xmin, rtol=0, atol=atol))
        + np.mean(np.isclose(x, xmax, rtol=0, atol=atol))
    )

    if abs(index_corr) > 0.995 and monotonic_fraction > 0.98:
        raise ValueError(
            f"{label} looks like a TIME/RAMP column, not acoustic voltage "
            f"(index correlation={index_corr:.6f}, monotonic fraction="
            f"{monotonic_fraction:.4f}). The script has stopped rather than "
            "producing misleading SPL graphs."
        )

    return {
        "index_correlation": index_corr,
        "monotonic_fraction": monotonic_fraction,
        "raw_min": xmin,
        "raw_max": xmax,
        "raw_std": float(np.std(x)),
        "endpoint_clip_fraction": clip_fraction,
    }



def detect_hard_flatline_tail(raw: np.ndarray) -> Tuple[int, Dict[str, float]]:
    """Detect a long, essentially constant tail caused by acquisition stopping.

    Returns the number of samples that should be kept. The detector only activates
    when the END of the record is orders of magnitude flatter than the active data.
    """
    x = np.asarray(raw, dtype=float)
    n = len(x)
    info = {
        "flatline_tail_detected": 0,
        "flatline_tail_start_sample": n,
        "flatline_tail_start_s": n / SAMPLE_RATE_HZ,
        "flatline_tail_samples_removed": 0,
    }
    if (not TRIM_HARD_FLATLINE_TAIL) or n < int(2*FLATLINE_MIN_TAIL_S*SAMPLE_RATE_HZ):
        return n, info

    finite = np.isfinite(x)
    if finite.sum() < 100:
        return n, info
    xf = x[finite]
    active_scale = float(np.std(xf))
    if not np.isfinite(active_scale) or active_scale <= 0:
        return n, info

    block = max(32, int(round(FLATLINE_BLOCK_S*SAMPLE_RATE_HZ)))
    min_tail = max(block, int(round(FLATLINE_MIN_TAIL_S*SAMPLE_RATE_HZ)))
    flat_std = max(1e-15, FLATLINE_RELATIVE_STD * active_scale)
    flat_ptp = max(1e-14, 10.0 * flat_std)

    # The final block must itself look hard-flat; otherwise do nothing.
    tail = x[max(0, n-block):n]
    tail = tail[np.isfinite(tail)]
    if len(tail) < block//2:
        return n, info
    if float(np.std(tail)) > flat_std or float(np.ptp(tail)) > flat_ptp:
        return n, info

    cutoff = n
    # Walk backwards in blocks while the data remain essentially constant.
    end = n
    while end > 0:
        start = max(0, end-block)
        b = x[start:end]
        b = b[np.isfinite(b)]
        if len(b) < max(8, (end-start)//2):
            break
        if float(np.std(b)) <= flat_std and float(np.ptp(b)) <= flat_ptp:
            cutoff = start
            end = start
        else:
            break

    removed = n - cutoff
    if removed < min_tail or cutoff < int(MIN_USABLE_ACOUSTIC_DURATION_S*SAMPLE_RATE_HZ):
        return n, info

    info.update({
        "flatline_tail_detected": 1,
        "flatline_tail_start_sample": int(cutoff),
        "flatline_tail_start_s": float(cutoff/SAMPLE_RATE_HZ),
        "flatline_tail_samples_removed": int(removed),
    })
    return int(cutoff), info


def safe_pressure_to_spl(p_rms):
    """SPL conversion that marks zero/invalid RMS windows as NaN, not -6000 dB."""
    p = np.asarray(p_rms, dtype=float)
    out = np.full_like(p, np.nan, dtype=float)
    m = np.isfinite(p) & (p > SPL_ZERO_FLOOR_PA)
    out[m] = 20.0*np.log10(p[m]/P_REF)
    if np.ndim(p_rms) == 0:
        return float(out)
    return out


def build_common_event_segments(signals: Dict[str, Tuple[np.ndarray, np.ndarray]]):
    """Build equal-duration, ignition-aligned segments for fair sensor comparisons."""
    onset = {}
    available = {}
    onset_info = {}
    for name, (t, p) in signals.items():
        o, info = acoustic_event_onset_from_pressure(p)
        onset[name] = float(o)
        onset_info[name] = info
        available[name] = len(p)/SAMPLE_RATE_HZ - o

    shortest = min(available.values())
    usable = shortest - COMMON_WINDOW_START_AFTER_IGNITION_S - COMMON_WINDOW_END_MARGIN_S
    if usable < MIN_USABLE_ACOUSTIC_DURATION_S:
        raise ValueError(
            f"Only {usable:.3f} s of common valid acoustic data remain after ignition; "
            f"need at least {MIN_USABLE_ACOUSTIC_DURATION_S:.3f} s."
        )

    n_common = int(math.floor(usable*SAMPLE_RATE_HZ))
    segments = {}
    rows = []
    for name, (t, p) in signals.items():
        start_s = onset[name] + COMMON_WINDOW_START_AFTER_IGNITION_S
        i0 = max(0, int(round(start_s*SAMPLE_RATE_HZ)))
        i1 = min(len(p), i0+n_common)
        # Enforce one exact length after rounding.
        nuse = i1-i0
        if nuse < n_common:
            n_common = nuse
        segments[name] = p[i0:i1]
        rows.append({
            "sensor": name,
            "detected_acoustic_onset_s": onset[name],
            "valid_time_after_onset_s": available[name],
            "comparison_start_after_onset_s": COMMON_WINDOW_START_AFTER_IGNITION_S,
            "comparison_duration_s": nuse/SAMPLE_RATE_HZ,
            **{f"onset_{k}": v for k, v in onset_info[name].items()},
        })

    # If rounding shortened one channel, enforce equal length everywhere.
    n_equal = min(len(v) for v in segments.values())
    segments = {k: v[:n_equal] for k, v in segments.items()}
    for row in rows:
        row["comparison_duration_s"] = n_equal/SAMPLE_RATE_HZ

    print(
        f"\nCommon ignition-aligned comparison window: "
        f"{COMMON_WINDOW_START_AFTER_IGNITION_S:.3f} to "
        f"{COMMON_WINDOW_START_AFTER_IGNITION_S+n_equal/SAMPLE_RATE_HZ:.3f} s "
        f"after acoustic onset ({n_equal:,} samples)."
    )
    return segments, pd.DataFrame(rows), onset

def remove_isolated_single_sample_spikes(p: np.ndarray) -> Tuple[np.ndarray, int]:
    """
    Conservative deglitcher: replace only isolated samples that create a very large
    jump immediately followed by an almost equal opposite jump.
    """
    x = np.asarray(p, dtype=float).copy()
    if not DESPIKE_ENABLED or len(x) < 5:
        return x, 0

    d = np.diff(x)
    med = np.median(d)
    mad = np.median(np.abs(d - med))
    robust_sigma = 1.4826 * mad
    if not np.isfinite(robust_sigma) or robust_sigma <= 0:
        return x, 0

    threshold = DESPIKE_THRESHOLD_MAD * robust_sigma
    d1 = d[:-1]
    d2 = d[1:]
    big_pair = (np.abs(d1) > threshold) & (np.abs(d2) > threshold)
    opposite = np.sign(d1) != np.sign(d2)
    cancellation = np.abs(d1 + d2) < 0.25*np.maximum(np.abs(d1), np.abs(d2))
    candidates = np.where(big_pair & opposite & cancellation)[0] + 1

    # Do not replace adjacent runs: those may be genuine transients.
    if len(candidates):
        keep = np.ones(len(candidates), dtype=bool)
        keep[1:] &= np.diff(candidates) > 1
        keep[:-1] &= np.diff(candidates) > 1
        candidates = candidates[keep]

    for i in candidates:
        x[i] = 0.5*(x[i-1] + x[i+1])
    return x, int(len(candidates))


def preprocess_acoustic_pressure(p: np.ndarray) -> Tuple[np.ndarray, Dict[str, float]]:
    x = np.asarray(p, dtype=float).copy()

    # Interpolate rare non-finite samples instead of silently replacing them with 0.
    finite = np.isfinite(x)
    nonfinite_count = int((~finite).sum())
    if not finite.all():
        good = np.where(finite)[0]
        if len(good) < 2:
            raise ValueError("Acoustic signal has insufficient finite data.")
        x[~finite] = np.interp(np.where(~finite)[0], good, x[good])

    if REMOVE_LINEAR_TREND:
        x = signal.detrend(x, type="linear")

    x, despiked = remove_isolated_single_sample_spikes(x)

    if APPLY_BANDPASS_FILTER:
        high = min(FILTER_HIGH_HZ, 0.95*(SAMPLE_RATE_HZ/2.0))
        low = max(FILTER_LOW_HZ, 0.1)
        if not (0 < low < high < SAMPLE_RATE_HZ/2):
            raise ValueError("Invalid band-pass filter limits.")
        sos = signal.butter(
            FILTER_ORDER, [low, high],
            btype="bandpass", fs=SAMPLE_RATE_HZ, output="sos"
        )
        x = signal.sosfiltfilt(sos, x)

    return x, {
        "nonfinite_samples_interpolated": nonfinite_count,
        "isolated_spikes_replaced": despiked,
    }


def savefig(fig, path: Path):
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def get_individual_plot_dir(outdir: Path) -> Path:
    """Return/create the subfolder containing per-sensor companion plots."""
    folder = outdir / "individual_sensor_plots"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


# ============================================================================
# FILE AUTO-DISCOVERY
# ============================================================================

def discover_input_files(folder: Path) -> Tuple[Path, Path]:
    supported = {".xlsx", ".xlsm", ".csv"}
    files = [
        p for p in folder.iterdir()
        if p.is_file()
        and p.suffix.lower() in supported
        and OUTPUT_FOLDER_NAME.lower() not in p.name.lower()
        and not p.name.startswith("~$")
    ]

    if TEST_1_FILE:
        t1 = folder / TEST_1_FILE
    else:
        t1 = find_named_test(files, 1)

    if TEST_2_FILE:
        t2 = folder / TEST_2_FILE
    else:
        t2 = find_named_test(files, 2)

    if t1 is None or t2 is None:
        # Last-resort: if exactly two supported files exist, use them in sorted order.
        candidates = sorted(files, key=lambda p: p.name.lower())
        if len(candidates) == 2:
            t1, t2 = candidates
        else:
            names = "\n".join(f"  - {p.name}" for p in candidates)
            raise FileNotFoundError(
                "Could not unambiguously identify Test 1 and Test 2 files.\n"
                "Put the script in the same folder as the two data files and name "
                "them something like 'Test 1.xlsx' and 'Test 2.xlsx'.\n\n"
                f"Supported files found:\n{names or '  (none)'}"
            )

    if t1.resolve() == t2.resolve():
        raise RuntimeError("Test 1 and Test 2 resolved to the same file.")

    return t1, t2


def find_named_test(files: List[Path], number: int) -> Optional[Path]:
    patterns = [
        rf"\btest[\s_-]*{number}\b",
        rf"\bdataset[\s_-]*{number}\b",
        rf"\bdata[\s_-]*{number}\b",
        rf"\btrial[\s_-]*{number}\b",
    ]
    hits = [
        p for p in files
        if any(re.search(pat, p.stem.lower()) for pat in patterns)
    ]
    if len(hits) == 1:
        return hits[0]
    if len(hits) > 1:
        # Prefer Excel because user stated the raw datasets are workbooks.
        hits = sorted(hits, key=lambda p: (p.suffix.lower() == ".csv", p.name.lower()))
        return hits[0]
    return None


# ============================================================================
# TABLE / WAVEFORM DETECTION
# ============================================================================

def detect_header_row(rows: List[List[object]]) -> Optional[Tuple[int, int, int, Optional[int]]]:
    """
    Return (header_row_index, ch1_col, ch2_col, time_col).
    """
    for ridx, row in enumerate(rows):
        ch1 = ch2 = time_col = None
        for cidx, value in enumerate(row):
            txt = normalise_header(value)
            n = channel_number_from_header(txt)
            if n == 1 and ch1 is None:
                ch1 = cidx
            elif n == 2 and ch2 is None:
                ch2 = cidx
            if is_time_header(txt) and time_col is None:
                time_col = cidx
        if ch1 is not None and ch2 is not None:
            return ridx, ch1, ch2, time_col
    return None


def rows_to_arrays(
    rows: Iterable[List[object]],
    ch1_idx: int,
    ch2_idx: int,
    time_idx: Optional[int],
    ch1_header: str,
    ch2_header: str,
    max_samples: int,
) -> Tuple[np.ndarray, np.ndarray, Optional[np.ndarray], float, float]:
    c1, c2, tt = [], [], []
    for row in rows:
        if len(c1) >= max_samples:
            break
        if ch1_idx >= len(row) or ch2_idx >= len(row):
            continue

        try:
            v1 = float(row[ch1_idx])
            v2 = float(row[ch2_idx])
        except (TypeError, ValueError):
            continue

        if not np.isfinite(v1) or not np.isfinite(v2):
            continue

        c1.append(v1)
        c2.append(v2)

        if time_idx is not None and time_idx < len(row):
            try:
                tv = float(row[time_idx])
            except (TypeError, ValueError):
                tv = np.nan
            tt.append(tv)

    time = np.asarray(tt, dtype=float) if time_idx is not None else None
    if time is not None and len(time) != len(c1):
        time = None

    return (
        np.asarray(c1, dtype=float),
        np.asarray(c2, dtype=float),
        time,
        voltage_scale_from_header(ch1_header),
        voltage_scale_from_header(ch2_header),
    )


def read_excel_waveform(path: Path, max_samples: int):
    if openpyxl is None:
        raise ImportError("openpyxl is required for Excel input: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_c1, all_c2, all_t = [], [], []
    scale1 = scale2 = 1.0
    found_any = False
    total = 0

    print(f"\nReading Excel workbook: {path.name}")
    print(f"Sheets: {wb.sheetnames}")

    for ws in wb.worksheets:
        if total >= max_samples:
            break

        # Scan the first 100 rows for a waveform header.
        preview = []
        row_iter = ws.iter_rows(values_only=True)
        for _ in range(100):
            try:
                preview.append(list(next(row_iter)))
            except StopIteration:
                break

        det = detect_header_row(preview)
        if det is None:
            print(f"  Skipping sheet '{ws.title}': no Channel 1/Channel 2 waveform header detected.")
            continue

        header_row, ch1_idx, ch2_idx, time_idx = det
        header = preview[header_row]
        h1 = str(header[ch1_idx])
        h2 = str(header[ch2_idx])
        s1 = voltage_scale_from_header(h1)
        s2 = voltage_scale_from_header(h2)

        # Include any data rows already consumed after the detected header.
        data_preview = preview[header_row + 1:]
        rows_remaining = max_samples - total

        c1a, c2a, ta, _, _ = rows_to_arrays(
            data_preview, ch1_idx, ch2_idx, time_idx, h1, h2, rows_remaining
        )

        remaining_after_preview = rows_remaining - len(c1a)
        c1b = c2b = np.array([], dtype=float)
        tb = None

        if remaining_after_preview > 0:
            # row_iter is already positioned after preview rows.
            c1b, c2b, tb, _, _ = rows_to_arrays(
                row_iter, ch1_idx, ch2_idx, time_idx, h1, h2, remaining_after_preview
            )

        c1 = np.concatenate([c1a, c1b])
        c2 = np.concatenate([c2a, c2b])

        if time_idx is not None and ta is not None and tb is not None:
            t = np.concatenate([ta, tb])
        elif time_idx is not None and ta is not None and len(c1b) == 0:
            t = ta
        elif time_idx is not None and tb is not None and len(c1a) == 0:
            t = tb
        else:
            t = None

        if len(c1) == 0:
            continue

        found_any = True
        scale1, scale2 = s1, s2
        all_c1.append(c1)
        all_c2.append(c2)
        if t is not None:
            all_t.append(t)
        else:
            all_t = []

        total += len(c1)
        print(
            f"  Using sheet '{ws.title}' — {len(c1):,} samples "
            f"(running total {total:,})"
        )

    wb.close()

    if not found_any:
        raise ValueError(
            f"{path.name} does not contain a detectable raw waveform table.\n"
            "The small example you supplied is a WaveForms MEASUREMENT SUMMARY "
            "(Extent/Name), not a raw waveform export. The real dataset needs "
            "numeric Channel 1 and Channel 2 sample columns."
        )

    c1 = np.concatenate(all_c1)[:max_samples]
    c2 = np.concatenate(all_c2)[:max_samples]
    t = None
    if all_t:
        t = np.concatenate(all_t)[:min(len(c1), sum(len(x) for x in all_t))]
        if len(t) != len(c1) or not np.all(np.isfinite(t)):
            t = None

    return c1, c2, t, scale1, scale2


def sniff_csv_delimiter(path: Path) -> str:
    with path.open("r", encoding="utf-8-sig", errors="replace") as f:
        sample = f.read(8192)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except Exception:
        return ","


def read_csv_waveform(path: Path, max_samples: int):
    """
    Read WaveForms-style CSV by exact detected column names.

    IMPORTANT: pandas returns integer usecols in FILE ORDER, not in the order given
    in usecols. The previous script assumed otherwise, so for a file ordered as
    Time, Channel 1, Channel 2 it analysed Time as CH1 and Channel 1 as CH2.
    This version explicitly selects each named column after loading.
    """
    delimiter = sniff_csv_delimiter(path)

    header_line_idx = None
    header_fields = None
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
        for idx, line in enumerate(f):
            if idx > 300:
                break
            if not line.strip() or line.lstrip().startswith("#"):
                continue
            try:
                row = next(csv.reader([line], delimiter=delimiter))
            except Exception:
                continue
            det = detect_header_row([row])
            if det is not None:
                header_line_idx = idx
                header_fields = row
                break

    if header_line_idx is None:
        raise ValueError(
            f"{path.name} does not contain a detectable raw waveform table. "
            "It needs numeric Time / Channel 1 / Channel 2 sample columns."
        )

    _, ch1_idx, ch2_idx, time_idx = detect_header_row([header_fields])
    ch1_name = header_fields[ch1_idx]
    ch2_name = header_fields[ch2_idx]
    time_name = header_fields[time_idx] if time_idx is not None else None

    # Read by names, then access by the same names. This is the crucial bug fix.
    wanted_names = [ch1_name, ch2_name] + ([time_name] if time_name is not None else [])
    df = pd.read_csv(
        path,
        sep=delimiter,
        skiprows=header_line_idx,
        header=0,
        usecols=lambda col: any(
            normalise_header(col) == normalise_header(w) for w in wanted_names
        ),
        nrows=max_samples,
        low_memory=False,
    )

    def resolve_loaded_column(target_name: str) -> str:
        matches = [
            c for c in df.columns
            if normalise_header(c) == normalise_header(target_name)
        ]
        if len(matches) != 1:
            raise ValueError(
                f"{path.name}: could not uniquely resolve column {target_name!r}. "
                f"Loaded columns: {list(df.columns)}"
            )
        return matches[0]

    c1_col = resolve_loaded_column(ch1_name)
    c2_col = resolve_loaded_column(ch2_name)
    t_col = resolve_loaded_column(time_name) if time_name is not None else None

    c1s = pd.to_numeric(df[c1_col], errors="coerce")
    c2s = pd.to_numeric(df[c2_col], errors="coerce")
    valid = c1s.notna() & c2s.notna()

    c1 = c1s.loc[valid].to_numpy(dtype=float)
    c2 = c2s.loc[valid].to_numpy(dtype=float)

    t = None
    if t_col is not None:
        ts = pd.to_numeric(df[t_col], errors="coerce").loc[valid]
        if ts.notna().all():
            t = ts.to_numpy(dtype=float)

    print(
        f"  CSV mapping confirmed: {ch1_name!r} -> CH1, "
        f"{ch2_name!r} -> CH2"
        + (f", {time_name!r} -> Time" if time_name is not None else "")
    )

    return (
        c1[:max_samples],
        c2[:max_samples],
        t[:max_samples] if t is not None else None,
        voltage_scale_from_header(ch1_name),
        voltage_scale_from_header(ch2_name),
    )


def read_waveform(path: Path, max_samples: int):
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel_waveform(path, max_samples)
    if suffix == ".csv":
        return read_csv_waveform(path, max_samples)
    raise ValueError(f"Unsupported input format: {path.suffix}")


# ============================================================================
# SIGNAL PROCESSING
# ============================================================================

def time_axis(t: Optional[np.ndarray], n: int) -> np.ndarray:
    # The experiment setting is authoritative: 250 kHz.
    # If a valid time vector exists, we use its starting time only and reconstruct a
    # uniform axis to prevent spreadsheet time-rounding noise from affecting analyses.
    start = 0.0
    if t is not None and len(t) and np.isfinite(t[0]):
        start = float(t[0])
    return start + np.arange(n, dtype=float) / SAMPLE_RATE_HZ


def sliding_spl(t: np.ndarray, p: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    nwin = max(2, int(round(SPL_WINDOW_S * SAMPLE_RATE_HZ)))
    step = max(1, int(round(SPL_STEP_S * SAMPLE_RATE_HZ)))
    if len(p) < nwin:
        return np.array([]), np.array([])

    x2 = np.nan_to_num(p, nan=0.0) ** 2
    cs = np.concatenate(([0.0], np.cumsum(x2)))
    starts = np.arange(0, len(p) - nwin + 1, step)
    sums = cs[starts + nwin] - cs[starts]
    rms = np.sqrt(sums / nwin)
    centers = starts + nwin // 2
    return t[centers], safe_pressure_to_spl(rms)


def sliding_spl_custom(
    t: np.ndarray,
    p: np.ndarray,
    window_s: float,
    step_s: float
) -> Tuple[np.ndarray, np.ndarray]:
    nwin = max(2, int(round(window_s * SAMPLE_RATE_HZ)))
    step = max(1, int(round(step_s * SAMPLE_RATE_HZ)))
    if len(p) < nwin:
        return np.array([]), np.array([])
    x2 = np.nan_to_num(p, nan=0.0) ** 2
    cs = np.concatenate(([0.0], np.cumsum(x2)))
    starts = np.arange(0, len(p) - nwin + 1, step)
    sums = cs[starts+nwin] - cs[starts]
    rms = np.sqrt(sums/nwin)
    centers = starts + nwin//2
    return t[centers], safe_pressure_to_spl(rms)


def welch_psd(p: np.ndarray):
    nper = min(WELCH_NPERSEG, len(p))
    nover = int(nper * WELCH_OVERLAP)
    return signal.welch(
        p,
        fs=SAMPLE_RATE_HZ,
        window="hann",
        nperseg=nper,
        noverlap=nover,
        detrend="constant",
        scaling="density",
    )


def a_weighting_db(f: np.ndarray) -> np.ndarray:
    f = np.asarray(f, dtype=float)
    out = np.full_like(f, -np.inf)
    mask = f > 0
    ff = f[mask]
    f2 = ff**2
    ra = (
        (12200.0**2) * (f2**2)
        / (
            (f2 + 20.6**2)
            * np.sqrt((f2 + 107.7**2) * (f2 + 737.9**2))
            * (f2 + 12200.0**2)
        )
    )
    out[mask] = 20*np.log10(ra) + 2.0
    return out


def integrate_psd_level(f, Pxx, weighted=False):
    mask = (f >= FREQ_MIN_HZ) & (f <= min(FREQ_MAX_HZ, f[-1]))
    ff = f[mask]
    pp = Pxx[mask].copy()
    if weighted:
        pp *= 10 ** (a_weighting_db(ff)/10.0)
    if len(ff) < 2:
        return np.nan
    return float(db10(np.trapezoid(pp, ff)/(P_REF**2)))


def third_octave_from_psd(f, Pxx):
    fmax = min(FREQ_MAX_HZ, float(f[-1]))
    kmin = math.floor(10*math.log10(max(FREQ_MIN_HZ, 1e-9)/1000.0)) - 1
    kmax = math.ceil(10*math.log10(fmax/1000.0)) + 1
    centers = 1000.0 * 10**(np.arange(kmin, kmax+1)/10.0)
    centers = centers[(centers >= FREQ_MIN_HZ) & (centers <= fmax)]
    edge = 2**(1/6)

    rows = []
    for fc in centers:
        lo, hi = fc/edge, fc*edge
        m = (f >= lo) & (f < hi)
        if m.sum() < 2:
            continue
        ms = np.trapezoid(Pxx[m], f[m])
        L = float(db10(ms/(P_REF**2)))
        A = float(a_weighting_db(np.array([fc]))[0])
        rows.append([fc, lo, hi, L, A, L+A])
    return pd.DataFrame(
        rows,
        columns=[
            "center_hz", "lower_hz", "upper_hz",
            "spl_db", "A_weighting_db", "A_weighted_spl_dbA"
        ]
    )


def average_psd(names, spectra):
    f0 = spectra[names[0]][0]
    powers = []
    for name in names:
        f, P = spectra[name]
        if np.array_equal(f, f0):
            powers.append(P)
        else:
            powers.append(np.interp(f0, f, P))
    return f0, np.mean(np.vstack(powers), axis=0)


def limited_cross_correlation(x, y):
    n = min(len(x), len(y))
    x = np.nan_to_num(x[:n] - np.nanmean(x[:n]))
    y = np.nan_to_num(y[:n] - np.nanmean(y[:n]))
    corr = signal.correlate(y, x, mode="full", method="fft")
    lags = signal.correlation_lags(len(y), len(x), mode="full")
    denom = np.sqrt(np.sum(x*x)*np.sum(y*y))
    if denom > 0:
        corr /= denom

    maxlag = int(MAX_CORRELATION_LAG_S*SAMPLE_RATE_HZ)
    m = np.abs(lags) <= maxlag
    lag_s = lags[m]/SAMPLE_RATE_HZ
    cc = corr[m]
    idx = int(np.argmax(np.abs(cc)))
    return lag_s, cc, float(lag_s[idx]), float(cc[idx])


def expected_window_correlation_peak(lag_s, cc, expected_s):
    m = np.abs(lag_s-expected_s) <= EXPECTED_DELAY_SEARCH_HALF_WIDTH_S
    if not np.any(m):
        return np.nan, np.nan
    ii = np.where(m)[0]
    j = ii[int(np.argmax(np.abs(cc[m])))]
    return float(lag_s[j]), float(cc[j])


def coherence_xy(x, y):
    n = min(len(x), len(y))
    nper = min(COHERENCE_NPERSEG, n)
    f, C = signal.coherence(
        x[:n], y[:n],
        fs=SAMPLE_RATE_HZ,
        window="hann",
        nperseg=nper,
        noverlap=nper//2,
    )
    return f, C



# ============================================================================
# STATIC-FIRE CHAMBER PRESSURE + EVENT ALIGNMENT
# ============================================================================

def _resolve_normalised_dataframe_column(df: pd.DataFrame, wanted: str) -> str:
    nw = normalise_header(wanted)
    matches = [c for c in df.columns if normalise_header(c) == nw]
    if not matches:
        raise KeyError(
            f"Could not find {wanted!r}. Available columns: {list(df.columns)}"
        )
    return matches[0]


def _first_sustained_true(mask: np.ndarray, min_samples: int) -> Optional[int]:
    if min_samples <= 1:
        idx = np.where(mask)[0]
        return int(idx[0]) if len(idx) else None
    run = np.convolve(mask.astype(int), np.ones(min_samples, dtype=int), mode="valid")
    idx = np.where(run >= min_samples)[0]
    return int(idx[0]) if len(idx) else None


def load_chamber_pressure(folder: Path) -> Dict[str, np.ndarray]:
    path = folder / CHAMBER_PRESSURE_FILE
    if not path.exists():
        warnings.warn(
            f"{CHAMBER_PRESSURE_FILE} was not found. Chamber-pressure alignment "
            "plots will be skipped."
        )
        return {}

    print(f"\nReading chamber pressure: {path.name} / sheet {CHAMBER_PRESSURE_SHEET!r}")
    df = pd.read_excel(
        path,
        sheet_name=CHAMBER_PRESSURE_SHEET,
        engine="openpyxl"
    )
    p_col = _resolve_normalised_dataframe_column(df, CHAMBER_PRESSURE_COLUMN)
    t_col = _resolve_normalised_dataframe_column(df, CHAMBER_TIME_COLUMN)

    t = pd.to_numeric(df[t_col], errors="coerce").to_numpy(dtype=float)
    p = pd.to_numeric(df[p_col], errors="coerce").to_numpy(dtype=float)
    valid = np.isfinite(t) & np.isfinite(p)
    t, p = t[valid], p[valid]
    order = np.argsort(t)
    t, p = t[order], p[order]

    if len(t) < 20:
        raise ValueError("Not enough chamber-pressure samples.")

    dt = float(np.median(np.diff(t)))
    fs = 1.0/dt

    # Robust smooth for onset detection only. Keep raw pressure for export.
    win = max(3, int(round(PRESSURE_SMOOTH_WINDOW_S*fs)))
    if win % 2 == 0:
        win += 1
    # median filter rejects the narrow ~100-unit spikes without deleting them from raw.
    p_smooth = signal.medfilt(p, kernel_size=win)

    baseline = float(np.median(p))
    mad = float(np.median(np.abs(p_smooth - baseline)))
    event_peak = float(np.max(p_smooth))
    dynamic = max(event_peak - baseline, EPS)
    threshold = baseline + max(8.0*1.4826*mad,
                               PRESSURE_ONSET_FRACTION_OF_EVENT*dynamic)

    sustain = max(1, int(round(PRESSURE_ONSET_SUSTAIN_S*fs)))
    onset_idx = _first_sustained_true(p_smooth > threshold, sustain)
    if onset_idx is None:
        onset_idx = int(np.argmax(np.gradient(p_smooth, t)))

    onset_t = float(t[onset_idx])
    t_rel = t - onset_t

    print(
        f"  DAQ rate ≈ {fs:.3f} Hz; detected static-fire onset at "
        f"DAQ t={onset_t:.3f} s; baseline={baseline:.4g}, "
        f"smoothed event peak={event_peak:.4g} {CHAMBER_PRESSURE_UNIT}"
    )

    return {
        "path": str(path),
        "t": t,
        "t_rel": t_rel,
        "p_raw": p,
        "p_smooth": p_smooth,
        "fs": fs,
        "baseline": baseline,
        "threshold": threshold,
        "onset_time_daq_s": onset_t,
    }


def acoustic_event_onset_from_pressure(p: np.ndarray) -> Tuple[float, Dict[str, float]]:
    """
    Detect the first strong acoustic event in one microphone record. Because the
    WaveForms and LabVIEW clocks are unrelated, this event is matched to the chamber
    pressure ignition event and becomes t=0 for source-event comparison.
    """
    x = np.asarray(p, dtype=float)
    nwin = max(4, int(round(ACOUSTIC_ONSET_ENVELOPE_WINDOW_S*SAMPLE_RATE_HZ)))
    # Efficient moving RMS, output at window centres.
    x2 = x*x
    cs = np.concatenate(([0.0], np.cumsum(x2)))
    starts = np.arange(0, len(x)-nwin+1)
    rms = np.sqrt((cs[starts+nwin]-cs[starts])/nwin)
    t_env = (starts + nwin//2)/SAMPLE_RATE_HZ

    # Smooth envelope slightly at the SPL time scale.
    smooth_n = max(1, int(round(0.010*SAMPLE_RATE_HZ)))
    # Downsample first to avoid huge convolutions.
    ds = max(1, int(round(0.001*SAMPLE_RATE_HZ)))  # 1 ms
    env_ds = rms[::ds]
    t_ds = t_env[::ds]
    if len(env_ds) < 5:
        return 0.0, {"threshold_pa_rms": np.nan, "detected": 0}

    # Low percentile estimates the noise floor even if the file starts at trigger.
    baseline = float(np.percentile(env_ds, 10))
    event_level = float(np.percentile(env_ds, 99.5))
    threshold = baseline + ACOUSTIC_ONSET_FRACTION_OF_EVENT*(event_level-baseline)

    sustain = max(1, int(round(ACOUSTIC_ONSET_SUSTAIN_S/0.001)))
    idx = _first_sustained_true(env_ds > threshold, sustain)
    if idx is None:
        idx = 0

    onset = float(t_ds[idx])
    # If capture begins already above threshold, the best supported statement is t=0.
    if idx <= 2:
        onset = 0.0

    return onset, {
        "threshold_pa_rms": threshold,
        "baseline_pa_rms": baseline,
        "event_level_pa_rms": event_level,
        "detected": 1,
    }


def plot_chamber_pressure_alignment(results, chamber, outdir, tables):
    if not chamber:
        return pd.DataFrame()

    alignment_rows = []
    for name, d in results.items():
        onset_s, onset_info = acoustic_event_onset_from_pressure(d["p"])
        align_t, align_spl = sliding_spl_custom(
            d["t"], d["p"], ALIGNMENT_SPL_WINDOW_S, ALIGNMENT_SPL_STEP_S
        )
        spl_rel = align_t - onset_s
        pressure_t = chamber["t_rel"]
        pressure_p = chamber["p_smooth"]

        # Only interpolate where the chamber data exist.
        p_interp = np.interp(
            spl_rel, pressure_t, pressure_p,
            left=np.nan, right=np.nan
        )

        aligned = pd.DataFrame({
            "event_relative_time_s": spl_rel,
            "spl_db": align_spl,
            "chamber_pressure": p_interp,
        })
        aligned.to_csv(
            tables/f"aligned_SPL_chamber_pressure_{name}.csv", index=False
        )

        fig, ax1 = plt.subplots(figsize=(11, 6))
        ax1.plot(spl_rel, align_spl, lw=1.2, label="Acoustic SPL (10 ms RMS)")
        ax1.set_xlabel("Event-relative time (s), ignition matched at t = 0")
        ax1.set_ylabel("SPL (dB re 20 µPa)")
        ax1.grid(True, alpha=0.3)

        ax2 = ax1.twinx()
        # Restrict pressure plotting to the acoustic overlap plus a small margin.
        if len(spl_rel):
            lo = max(float(np.nanmin(spl_rel))-0.15, float(np.min(pressure_t)))
            hi = min(float(np.nanmax(spl_rel))+0.15, float(np.max(pressure_t)))
            pm = (pressure_t >= lo) & (pressure_t <= hi)
        else:
            pm = np.ones_like(pressure_t, dtype=bool)
        ax2.plot(
            pressure_t[pm], pressure_p[pm],
            lw=1.4, ls="--", label="Combustion-chamber pressure"
        )
        ax2.set_ylabel(f"Chamber pressure ({CHAMBER_PRESSURE_UNIT})")

        ax1.axvline(0.0, lw=1.0, ls=":", alpha=0.7)
        travel_ms = 1000.0*DISTANCES_M[name]/SPEED_OF_SOUND_M_S
        ax1.set_title(
            f"18. SPL vs Combustion-Chamber Pressure — {DISPLAY_NAMES[name]}\n"
            f"Event-synchronised (arrival delay removed); nominal travel = {travel_ms:.1f} ms"
        )

        lines = ax1.get_lines()[:1] + ax2.get_lines()[:1]
        labels = [ln.get_label() for ln in lines]
        ax1.legend(lines, labels, loc="best")
        savefig(fig, outdir/f"18_SPL_vs_chamber_pressure_{name}.png")

        alignment_rows.append({
            "sensor": name,
            "distance_m": DISTANCES_M[name],
            "acoustic_onset_in_record_s": onset_s,
            "nominal_sound_travel_time_s": DISTANCES_M[name]/SPEED_OF_SOUND_M_S,
            "nominal_sound_travel_time_ms":
                1000.0*DISTANCES_M[name]/SPEED_OF_SOUND_M_S,
            "daq_pressure_onset_original_time_s": chamber["onset_time_daq_s"],
            "pressure_alignment_method":
                "ignition-event matching; absolute clocks not shared",
            **onset_info,
        })

    df = pd.DataFrame(alignment_rows)
    df.to_csv(tables/"chamber_pressure_alignment_summary.csv", index=False)

    # Save chamber pressure around the event for easy inspection.
    m = (chamber["t_rel"] >= -1.0) & (chamber["t_rel"] <= 12.0)
    pd.DataFrame({
        "event_relative_time_s": chamber["t_rel"][m],
        "chamber_pressure_raw": chamber["p_raw"][m],
        "chamber_pressure_smoothed_for_alignment": chamber["p_smooth"][m],
    }).to_csv(tables/"chamber_pressure_event.csv", index=False)

    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(
        chamber["t_rel"][m], chamber["p_raw"][m],
        lw=0.7, alpha=0.5, label="Raw chamber pressure"
    )
    ax.plot(
        chamber["t_rel"][m], chamber["p_smooth"][m],
        lw=1.5, label="Smoothed alignment trace"
    )
    ax.axvline(0, ls=":", label="Detected ignition")
    ax.set_xlabel("Event-relative time (s)")
    ax.set_ylabel(f"Chamber pressure ({CHAMBER_PRESSURE_UNIT})")
    ax.set_title("Static-Fire Combustion-Chamber Pressure Used for Alignment")
    ax.grid(True, alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"18a_chamber_pressure_alignment_reference.png")

    return df


def single_sided_fft_rms(p: np.ndarray):
    """Return frequency and Hann-window-corrected single-sided FFT RMS amplitude.

    Units are Pa RMS per FFT bin. This is a literal FFT-amplitude product; for
    broadband noise, PSD/Welch remains the preferred record-length-independent plot.
    """
    x = np.asarray(p, dtype=float)
    x = x[np.isfinite(x)]
    if len(x) < 4:
        return np.array([]), np.array([])
    x = signal.detrend(x, type="constant")
    n = len(x)
    w = signal.windows.hann(n, sym=False)
    cg = np.sum(w) / n
    X = np.fft.rfft(x*w)
    amp_peak = np.abs(X) / (n*max(cg, EPS))
    if n > 1:
        amp_peak[1:-1] *= 2.0
        if n % 2:  # odd-length rfft has no Nyquist bin
            amp_peak[-1] *= 2.0
    amp_rms = amp_peak / np.sqrt(2.0)
    # DC is not meaningful after detrending and is excluded in plots anyway.
    f = np.fft.rfftfreq(n, d=1.0/SAMPLE_RATE_HZ)
    return f, amp_rms


def plot_fft_all_sensors(results, outdir, tables):
    """
    Plot and export a direct single-sided FFT.

    Creates:
      * one combined FFT containing every sensor;
      * one individual FFT for each sensor;
      * one merged CSV plus one CSV per sensor.
    """
    individual_dir = get_individual_plot_dir(outdir)
    merged = None

    # Combined figure
    fig, ax = plt.subplots(figsize=(11, 7))

    for name, d in results.items():
        f, amp = single_sided_fft_rms(d["p"])
        if len(f) == 0:
            continue

        m = (f >= FFT_MIN_HZ) & (f <= min(FFT_MAX_HZ, f[-1]))
        ff = f[m]
        aa = amp[m]
        level = safe_pressure_to_spl(aa)

        # Add sensor to combined graph
        ax.semilogx(ff, level, lw=0.9, label=DISPLAY_NAMES[name])

        # Export this sensor's FFT numerical data
        one = pd.DataFrame({
            "frequency_hz": ff,
            "fft_rms_pa_per_bin": aa,
            "fft_level_db_re_20uPa_per_bin": level,
        })
        one.to_csv(tables/f"fft_amplitude_{name}.csv", index=False)

        # Individual graph
        fig_ind, ax_ind = plt.subplots(figsize=(11, 7))
        ax_ind.semilogx(ff, level, lw=1.0)
        ax_ind.set_xlim(FFT_MIN_HZ, FFT_MAX_HZ)
        ax_ind.set_xlabel("Frequency (Hz)")
        ax_ind.set_ylabel("FFT amplitude level (dB re 20 µPa RMS per bin)")
        ax_ind.set_title(
            f"20. Single-Sided FFT Amplitude — {DISPLAY_NAMES[name]}\n"
            "Common ignition-aligned analysis window"
        )
        ax_ind.grid(True, which="both", alpha=0.3)
        savefig(fig_ind, individual_dir/f"20_fft_{name}.png")

        # Merge into all-sensor CSV
        merged_one = pd.DataFrame({
            "frequency_hz": ff,
            f"{name}_fft_rms_pa_per_bin": aa,
            f"{name}_fft_level_db_re_20uPa_per_bin": level,
        })
        if merged is None:
            merged = merged_one
        else:
            merged = pd.merge(merged, merged_one, on="frequency_hz", how="outer")

    # Finish combined graph
    ax.set_xlim(FFT_MIN_HZ, FFT_MAX_HZ)
    ax.set_xlabel("Frequency (Hz)")
    ax.set_ylabel("FFT amplitude level (dB re 20 µPa RMS per bin)")
    ax.set_title(
        "20. Single-Sided FFT Amplitude — All Sensors\n"
        "Common ignition-aligned window; use PSD for broadband level comparisons"
    )
    ax.grid(True, which="both", alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"20_fft_amplitude_all_sensors.png")

    if merged is not None:
        merged.sort_values("frequency_hz").to_csv(
            tables/"fft_amplitude_all_sensors.csv",
            index=False
        )

    return merged if merged is not None else pd.DataFrame()



def build_sensor_health_summary(qc_df: pd.DataFrame, pair_df: pd.DataFrame, metrics: pd.DataFrame):
    """Combine acquisition, pair-delay/coherence, and within-pair level checks."""
    q = qc_df.set_index("sensor", drop=False)
    m = metrics.set_index("sensor", drop=False)
    rows = []
    pair_lookup = {}
    for _, pr in pair_df.iterrows():
        for ch in (pr["sensor_a"], pr["sensor_b"]):
            pair_lookup[ch] = pr

    for ch in q.index:
        flags = []
        severity = "PASS"
        qr = q.loc[ch]
        if int(qr.get("flatline_tail_detected", 0)):
            flags.append("hard acquisition dropout / partial record")
            severity = "FAIL"
        ratio = float(qr.get("absolute_to_robust_peak_ratio", np.nan))
        if np.isfinite(ratio) and ratio > 2.0:
            flags.append("isolated impulsive peak(s)")
            if severity == "PASS": severity = "CHECK"

        pr = pair_lookup.get(ch)
        pair_level_diff = np.nan
        delay_error = np.nan
        median_coh = np.nan
        if pr is not None:
            a, b = pr["sensor_a"], pr["sensor_b"]
            if a in m.index and b in m.index:
                La, Lb = float(m.loc[a,"rms_spl_db"]), float(m.loc[b,"rms_spl_db"])
                pair_level_diff = abs(La-Lb)
                if pair_level_diff > PAIR_LEVEL_MISMATCH_WARN_DB:
                    flags.append(f"1 m pair broadband level mismatch = {pair_level_diff:.1f} dB")
                    if severity == "PASS": severity = "CHECK"
            expd = float(pr.get("expected_signed_delay_ms", np.nan))
            measd = float(pr.get("peak_near_expected_lag_ms", np.nan))
            if np.isfinite(expd) and np.isfinite(measd):
                delay_error = abs(measd-expd)
                if delay_error > PAIR_DELAY_ERROR_WARN_MS:
                    flags.append(f"pair delay error = {delay_error:.2f} ms")
                    if severity == "PASS": severity = "CHECK"
            median_coh = float(pr.get("median_coherence_80_to_1000_Hz", np.nan))
            if (not np.isfinite(median_coh)) or median_coh < PAIR_COHERENCE_WARN_BELOW:
                flags.append(f"low 80–1000 Hz pair coherence = {median_coh:.2f}" if np.isfinite(median_coh) else "pair coherence unavailable")
                if severity == "PASS": severity = "CHECK"

        if not flags:
            flags = ["no automatic QC fault detected"]
        rows.append({
            "sensor": ch,
            "description": DISPLAY_NAMES[ch],
            "health": severity,
            "valid_duration_s": float(qr.get("valid_duration_s", np.nan)),
            "coverage_fraction": float(qr.get("coverage_fraction", np.nan)),
            "pair_rms_level_difference_db": pair_level_diff,
            "pair_delay_error_ms": delay_error,
            "pair_median_coherence_80_1000_hz": median_coh,
            "notes": "; ".join(flags),
        })
    return pd.DataFrame(rows)


# ============================================================================
# PLOTS
# ============================================================================

def plot_pressure_time(results, outdir):
    individual_dir = get_individual_plot_dir(outdir)

    # Combined graph
    fig, ax = plt.subplots(figsize=(12, 6))
    for name, d in results.items():
        stride = max(1, math.ceil(len(d["p"])/120_000))
        ax.plot(
            d["t"][::stride]-d["t"][0],
            d["p"][::stride],
            lw=0.8,
            label=DISPLAY_NAMES[name]
        )
    ax.set_xlabel("Time from start of analysed record (s)")
    ax.set_ylabel("Calibrated acoustic pressure (Pa)")
    ax.set_title("1. Filtered Calibrated Acoustic Pressure vs Time — All Sensors")
    ax.grid(True, alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"01_calibrated_pressure_vs_time.png")

    # Individual companion graphs
    for name, d in results.items():
        stride = max(1, math.ceil(len(d["p"])/120_000))
        fig, ax = plt.subplots(figsize=(11, 6))
        ax.plot(
            d["t"][::stride]-d["t"][0],
            d["p"][::stride],
            lw=0.9
        )
        ax.set_xlabel("Time from start of analysed record (s)")
        ax.set_ylabel("Calibrated acoustic pressure (Pa)")
        ax.set_title(f"1. Calibrated Acoustic Pressure vs Time — {DISPLAY_NAMES[name]}")
        ax.grid(True, alpha=0.3)
        savefig(fig, individual_dir/f"01_pressure_time_{name}.png")


def plot_spl_time(results, outdir):
    individual_dir = get_individual_plot_dir(outdir)

    # Combined graph
    fig, ax = plt.subplots(figsize=(12, 6))
    for name, d in results.items():
        ax.plot(
            d["spl_t"]-d["t"][0],
            d["spl_db"],
            label=DISPLAY_NAMES[name]
        )
    ax.set_xlabel("Time from start of analysed record (s)")
    ax.set_ylabel("SPL (dB re 20 µPa)")
    ax.set_title("2. SPL vs Time — All Sensors")
    ax.grid(True, alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"02_spl_vs_time.png")

    # Individual companion graphs
    for name, d in results.items():
        fig, ax = plt.subplots(figsize=(11, 6))
        ax.plot(d["spl_t"]-d["t"][0], d["spl_db"], lw=1.0)
        ax.set_xlabel("Time from start of analysed record (s)")
        ax.set_ylabel("SPL (dB re 20 µPa)")
        ax.set_title(f"2. SPL vs Time — {DISPLAY_NAMES[name]}")
        ax.grid(True, alpha=0.3)
        savefig(fig, individual_dir/f"02_spl_time_{name}.png")


def plot_distance_metric(metrics, col, title, filename, outdir):
    m = metrics.sort_values("distance_m")
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(m["distance_m"], m[col], "o-")
    for _, row in m.iterrows():
        ax.annotate(
            row["sensor"],
            (row["distance_m"], row[col]),
            xytext=(5, 5),
            textcoords="offset points",
            fontsize=8
        )
    ax.set_xlabel("Distance (m)")
    ax.set_ylabel("SPL (dB re 20 µPa)")
    ax.set_title(title)
    ax.grid(True, alpha=0.3)
    savefig(fig, outdir/filename)


def plot_attenuation(metrics, outdir):
    m = metrics.sort_values("distance_m").reset_index(drop=True)
    r0 = m.loc[0, "distance_m"]
    L0 = m.loc[0, "rms_spl_db"]
    measured = L0 - m["rms_spl_db"].to_numpy()
    theoretical = 20*np.log10(m["distance_m"].to_numpy()/r0)

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(m["distance_m"], measured, "o-", label="Measured")
    ax.plot(
        m["distance_m"], theoretical, "s--",
        label=r"Theoretical $20\log_{10}(r/r_0)$"
    )
    ax.set_xlabel("Distance (m)")
    ax.set_ylabel("Attenuation relative to closest sensor (dB)")
    ax.set_title("5. Measured Attenuation vs Theory — Common Ignition-Aligned Window")
    ax.grid(True, alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"05_measured_vs_theoretical_attenuation.png")


def plot_all_psd(results, spectra, outdir):
    individual_dir = get_individual_plot_dir(outdir)

    # Combined graph
    fig, ax = plt.subplots(figsize=(10, 6))
    for name, (f, P) in spectra.items():
        ax.semilogx(f[1:], db10(P[1:]/P_REF**2), label=DISPLAY_NAMES[name])
    ax.set_xlim(FREQ_MIN_HZ, FREQ_MAX_HZ)
    ax.set_xlabel("Frequency (Hz)")
    ax.set_ylabel("PSD level (dB re (20 µPa)²/Hz)")
    ax.set_title("6. PSD — All Sensors (Common Ignition-Aligned Window)")
    ax.grid(True, which="both", alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"06_psd_all_sensors.png")

    # Individual companion graphs
    for name, (f, P) in spectra.items():
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.semilogx(f[1:], db10(P[1:]/P_REF**2), lw=1.0)
        ax.set_xlim(FREQ_MIN_HZ, FREQ_MAX_HZ)
        ax.set_xlabel("Frequency (Hz)")
        ax.set_ylabel("PSD level (dB re (20 µPa)²/Hz)")
        ax.set_title(f"6. PSD — {DISPLAY_NAMES[name]}")
        ax.grid(True, which="both", alpha=0.3)
        savefig(fig, individual_dir/f"06_psd_{name}.png")


def plot_near_far_psd(spectra, outdir):
    fn, Pn = average_psd(NEAR_SENSORS, spectra)
    ff, Pf = average_psd(FAR_SENSORS, spectra)
    if not np.array_equal(fn, ff):
        Pf = np.interp(fn, ff, Pf)
    rn = np.mean([DISTANCES_M[x] for x in NEAR_SENSORS])
    rf = np.mean([DISTANCES_M[x] for x in FAR_SENSORS])

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.semilogx(fn[1:], db10(Pn[1:]/P_REF**2), label=f"Near average — {rn:.1f} m")
    ax.semilogx(fn[1:], db10(Pf[1:]/P_REF**2), label=f"Far average — {rf:.1f} m")
    ax.set_xlim(FREQ_MIN_HZ, FREQ_MAX_HZ)
    ax.set_xlabel("Frequency (Hz)")
    ax.set_ylabel("PSD level (dB re (20 µPa)²/Hz)")
    ax.set_title("7. Near-Array vs Far-Array Average PSD — Common Window")
    ax.grid(True, which="both", alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"07_near_vs_far_average_psd.png")

    return fn, Pn, Pf, rn, rf


def plot_spectral_attenuation(f, Pn, Pf, rn, rf, outdir):
    attenuation = db10(Pn/np.maximum(Pf, EPS))
    theoretical = 20*np.log10(rf/rn)

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.semilogx(f[1:], attenuation[1:], label="Measured spectral attenuation")
    ax.axhline(
        theoretical, ls="--",
        label=f"Theoretical 20log₁₀({rf:.1f}/{rn:.1f}) = {theoretical:.2f} dB"
    )
    ax.axhline(0.0, ls=":", alpha=0.6, label="0 dB (near = far)")
    ax.set_xlim(FREQ_MIN_HZ, FREQ_MAX_HZ)
    ax.set_xlabel("Frequency (Hz)")
    ax.set_ylabel("Near − far level (dB)")
    ax.set_title(f"8. Spectral Attenuation: {rn:.1f} m → {rf:.1f} m")
    ax.grid(True, which="both", alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"08_spectral_attenuation_23p3m_to_69p0m.png")


def plot_third_octave_near_far(f, Pn, Pf, rn, rf, outdir, tables):
    ndf = third_octave_from_psd(f, Pn)
    fdf = third_octave_from_psd(f, Pf)
    ndf.to_csv(tables/"near_average_third_octave.csv", index=False)
    fdf.to_csv(tables/"far_average_third_octave.csv", index=False)

    merged = pd.merge(ndf, fdf, on="center_hz", suffixes=("_near", "_far"))
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.semilogx(
        merged["center_hz"], merged["spl_db_near"],
        "o-", label=f"Near average — {rn:.1f} m"
    )
    ax.semilogx(
        merged["center_hz"], merged["spl_db_far"],
        "s-", label=f"Far average — {rf:.1f} m"
    )
    ax.set_xlabel("1/3-octave centre frequency (Hz)")
    ax.set_ylabel("Band SPL (dB re 20 µPa)")
    ax.set_title("9. 1/3-Octave SPL — Near vs Far — Common Window")
    ax.grid(True, which="both", alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"09_third_octave_near_vs_far.png")


def plot_spectrograms(results, outdir):
    for name, d in results.items():
        p = d["p"]
        nper = min(SPECTROGRAM_NPERSEG, len(p))
        nover = int(nper*SPECTROGRAM_OVERLAP)
        f, t, S = signal.spectrogram(
            p,
            fs=SAMPLE_RATE_HZ,
            window="hann",
            nperseg=nper,
            noverlap=nover,
            scaling="density",
            mode="psd",
        )
        m = (f >= FREQ_MIN_HZ) & (f <= FREQ_MAX_HZ)
        Z = db10(S[m, :]/P_REF**2)

        # Display floor only: prevents zeros/near-zeros from forcing a -3000 dB
        # colour scale and making the useful data unreadable.
        finite = Z[np.isfinite(Z)]
        vmax = float(np.percentile(finite, 99.5)) if finite.size else 0.0
        vmin = vmax - SPECTROGRAM_DISPLAY_RANGE_DB

        fig, ax = plt.subplots(figsize=(11, 6))
        pcm = ax.pcolormesh(
            t, f[m], Z,
            shading="auto", vmin=vmin, vmax=vmax
        )
        ax.set_yscale("log")
        ax.set_xlabel("Time (s)")
        ax.set_ylabel("Frequency (Hz)")
        ax.set_title(f"10. Spectrogram — {DISPLAY_NAMES[name]}")
        fig.colorbar(pcm, ax=ax, label="PSD level (dB re (20 µPa)²/Hz)")
        savefig(fig, outdir/f"10_spectrogram_{name}.png")


def plot_pairs(results, outdir, tables):
    rows = []
    for idx, (pair_name, a, b) in enumerate(PAIR_DEFINITIONS, start=1):
        lag, cc, global_lag, global_corr = limited_cross_correlation(
            results[a]["p"], results[b]["p"]
        )

        signed_distance = DISTANCES_M[b] - DISTANCES_M[a]
        expected = signed_distance/SPEED_OF_SOUND_M_S
        local_lag, local_corr = expected_window_correlation_peak(lag, cc, expected)

        fig, ax = plt.subplots(figsize=(9, 5))
        ax.plot(lag*1000, cc)
        ax.axvline(
            expected*1000, ls=":",
            label=f"Expected Δr/c = {expected*1000:.3f} ms"
        )
        if np.isfinite(local_lag):
            ax.axvline(
                local_lag*1000, ls="--",
                label=f"Peak near expected = {local_lag*1000:.3f} ms"
            )
        if abs(global_lag-local_lag) > 1e-9 if np.isfinite(local_lag) else True:
            ax.axvline(global_lag*1000, ls="-.", alpha=0.7,
                       label=f"Global |corr| peak = {global_lag*1000:.3f} ms")
        ax.set_xlabel("Lag (ms)")
        ax.set_ylabel("Normalized cross-correlation")
        title_number = 10 + idx
        ax.set_title(f"{title_number}. Cross-Correlation — {pair_name}")
        ax.grid(True, alpha=0.3)
        ax.legend()
        savefig(
            fig,
            outdir/f"{title_number:02d}_cross_correlation_{safe_name(pair_name)}.png"
        )

        f, coh = coherence_xy(results[a]["p"], results[b]["p"])
        band = (f >= PAIR_COHERENCE_BAND_HZ[0]) & (f <= PAIR_COHERENCE_BAND_HZ[1])
        median_coh = float(np.nanmedian(coh[band])) if np.any(band) else np.nan
        pair_status = "OK"
        if (not np.isfinite(median_coh)) or median_coh < PAIR_COHERENCE_WARN_BELOW:
            pair_status = "CHECK — LOW PAIR COHERENCE"
        if np.isfinite(local_corr) and local_corr < -0.2:
            pair_status += "; POSSIBLE POLARITY / CHANNEL MISMATCH"

        fig, ax = plt.subplots(figsize=(9, 5))
        ax.semilogx(f[1:], coh[1:])
        ax.axhline(PAIR_COHERENCE_WARN_BELOW, ls="--", alpha=0.6,
                   label=f"QC reference = {PAIR_COHERENCE_WARN_BELOW:.2f}")
        ax.set_xlim(FREQ_MIN_HZ, FREQ_MAX_HZ)
        ax.set_ylim(0, 1.02)
        ax.set_xlabel("Frequency (Hz)")
        ax.set_ylabel("Magnitude-squared coherence")
        ax.set_title(
            f"13. Coherence — {pair_name}\n"
            f"median {PAIR_COHERENCE_BAND_HZ[0]:.0f}–{PAIR_COHERENCE_BAND_HZ[1]:.0f} Hz = {median_coh:.2f}"
        )
        ax.grid(True, which="both", alpha=0.3)
        ax.legend()
        savefig(fig, outdir/f"13_coherence_{safe_name(pair_name)}.png")

        rows.append({
            "pair": pair_name,
            "sensor_a": a,
            "sensor_b": b,
            "signed_distance_b_minus_a_m": signed_distance,
            "expected_signed_delay_ms": expected*1000,
            "peak_near_expected_lag_ms": local_lag*1000 if np.isfinite(local_lag) else np.nan,
            "peak_near_expected_correlation": local_corr,
            "global_abs_peak_lag_ms": global_lag*1000,
            "global_abs_peak_correlation": global_corr,
            "median_coherence_80_to_1000_Hz": median_coh,
            "pair_status": pair_status,
        })

    df = pd.DataFrame(rows)
    df.to_csv(tables/"pair_analysis.csv", index=False)
    return df

def plot_distance_corrected(spectra, outdir):
    individual_dir = get_individual_plot_dir(outdir)

    # Combined graph
    fig, ax = plt.subplots(figsize=(10, 6))
    for name, (f, P) in spectra.items():
        r = DISTANCES_M[name]
        level = db10(P/P_REF**2)
        corrected = level + 20*np.log10(r/DISTANCE_REFERENCE_M)
        ax.semilogx(
            f[1:], corrected[1:],
            label=f"{DISPLAY_NAMES[name]} → {DISTANCE_REFERENCE_M:g} m"
        )
    ax.set_xlim(FREQ_MIN_HZ, FREQ_MAX_HZ)
    ax.set_xlabel("Frequency (Hz)")
    ax.set_ylabel("Distance-corrected PSD level (dB/Hz)")
    ax.set_title("14. Distance-Corrected Spectra — All Sensors")
    ax.grid(True, which="both", alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"14_distance_corrected_spectra.png")

    # Individual companion graphs
    for name, (f, P) in spectra.items():
        r = DISTANCES_M[name]
        level = db10(P/P_REF**2)
        corrected = level + 20*np.log10(r/DISTANCE_REFERENCE_M)
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.semilogx(f[1:], corrected[1:], lw=1.0)
        ax.set_xlim(FREQ_MIN_HZ, FREQ_MAX_HZ)
        ax.set_xlabel("Frequency (Hz)")
        ax.set_ylabel("Distance-corrected PSD level (dB/Hz)")
        ax.set_title(
            f"14. Distance-Corrected Spectrum — {DISPLAY_NAMES[name]} "
            f"→ {DISTANCE_REFERENCE_M:g} m"
        )
        ax.grid(True, which="both", alpha=0.3)
        savefig(fig, individual_dir/f"14_distance_corrected_{name}.png")


def choose_source_estimate_channels(qc_df: pd.DataFrame, pair_df: pd.DataFrame):
    """Select the most trustworthy channels for the 1 m source-equivalent estimate.

    Priority is given to a simultaneous microphone pair that:
      * has no hard acquisition dropout (unless explicitly allowed),
      * has reasonable pair coherence, and
      * has a measured delay close to the geometric 1 m propagation delay.

    This deliberately prevents a broken / zero-filled near recording from being
    averaged into the source estimate.
    """
    qc = qc_df.set_index("sensor", drop=False)
    candidates = []
    reasons = {}

    for _, row in pair_df.iterrows():
        a, b = row["sensor_a"], row["sensor_b"]
        pair_ok = True
        why = []

        coh = row.get("median_coherence_80_to_1000_Hz", np.nan)
        if (not np.isfinite(coh)) or float(coh) < SOURCE_ESTIMATE_MIN_COHERENCE:
            pair_ok = False
            why.append(f"median coherence {coh:.2f} < {SOURCE_ESTIMATE_MIN_COHERENCE:.2f}" if np.isfinite(coh) else "coherence unavailable")

        expected = row.get("expected_signed_delay_ms", np.nan)
        measured = row.get("peak_near_expected_lag_ms", np.nan)
        if np.isfinite(expected) and np.isfinite(measured):
            err = abs(float(measured)-float(expected))
            if err > SOURCE_ESTIMATE_MAX_DELAY_ERROR_MS:
                pair_ok = False
                why.append(f"delay error {err:.2f} ms")

        for ch in (a, b):
            if ch not in qc.index:
                pair_ok = False
                why.append(f"{ch} missing QC")
                continue
            if SOURCE_ESTIMATE_REQUIRE_FULL_RECORD:
                if int(qc.loc[ch, "flatline_tail_detected"]) != 0 or float(qc.loc[ch, "coverage_fraction"]) < 0.95:
                    pair_ok = False
                    why.append(f"{ch} partial/dropout")

        if pair_ok:
            # Score by coherence, best first.
            candidates.append((float(coh), [a, b], row["pair"]))
        else:
            reasons[str(row["pair"])] = "; ".join(why) if why else "failed QC"

    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        _, selected, selected_pair = candidates[0]
        return selected, f"Selected {selected_pair} by dropout/coherence/delay QC", reasons

    # Fallback: use all channels with full records, but make the weaker basis clear.
    selected = []
    for ch, row in qc.iterrows():
        full = int(row["flatline_tail_detected"]) == 0 and float(row["coverage_fraction"]) >= 0.95
        if full:
            selected.append(ch)
    if selected:
        return selected, "No pair passed all source-estimate QC; using full-record channels only", reasons

    # Last fallback: all available channels. This should be rare and is explicitly labelled.
    return list(qc.index), "WARNING: no full-record channels passed QC; using all available channels", reasons


def plot_source_equivalent_spl(full_results, qc_df, pair_df, outdir, tables):
    """Back-propagate measured SPL to a 1 m free-field equivalent.

    L_1m(t) = L_r(t) + 20 log10(r / 1 m)

    This is intentionally called *source-equivalent*, not SPL at r=0.  SPL at an
    ideal point source is undefined/singular, and a rocket plume is distributed and
    strongly directional.  The plot is therefore a rough engineering estimate.
    """
    selected, selection_note, rejection_reasons = choose_source_estimate_channels(qc_df, pair_df)

    traces = {}
    rows = []
    for name, d in full_results.items():
        onset_s, _ = acoustic_event_onset_from_pressure(d["p"])
        st, sdb = sliding_spl_custom(
            d["t"], d["p"],
            SOURCE_ESTIMATE_SPL_WINDOW_S,
            SOURCE_ESTIMATE_SPL_STEP_S,
        )
        if len(st) == 0:
            continue
        rel = st - (d["t"][0] + onset_s)
        correction = 20.0*np.log10(DISTANCES_M[name]/SOURCE_EQUIVALENT_REFERENCE_M)
        src = sdb + correction
        valid = np.isfinite(rel) & np.isfinite(src)
        rel, src = rel[valid], src[valid]
        traces[name] = (rel, src, correction)
        for tt, ll in zip(rel, src):
            rows.append({
                "sensor": name,
                "event_relative_time_s": tt,
                "measured_distance_m": DISTANCES_M[name],
                "distance_correction_db": correction,
                "source_equivalent_SPL_at_1m_db": ll,
                "included_in_consensus": int(name in selected),
            })

    if not traces:
        return pd.DataFrame(), selected, selection_note

    # Common event-relative time grid across the selected channels.
    selected_available = [ch for ch in selected if ch in traces]
    if not selected_available:
        selected_available = list(traces)
        selection_note += "; selected channels lacked valid SPL history, so all valid traces were used"

    lo = max(float(np.nanmin(traces[ch][0])) for ch in selected_available)
    hi = min(float(np.nanmax(traces[ch][0])) for ch in selected_available)
    step = SOURCE_ESTIMATE_SPL_STEP_S
    if hi <= lo + step:
        consensus_t = np.array([])
        consensus = np.array([])
        p10 = p90 = np.array([])
    else:
        consensus_t = np.arange(lo, hi + 0.5*step, step)
        stack = []
        for ch in selected_available:
            tt, ll, _ = traces[ch]
            stack.append(np.interp(consensus_t, tt, ll))
        stack = np.vstack(stack)
        # Median in dB is robust to residual per-channel calibration mismatch.
        consensus = np.nanmedian(stack, axis=0)
        p10 = np.nanpercentile(stack, 10, axis=0)
        p90 = np.nanpercentile(stack, 90, axis=0)

    fig, ax = plt.subplots(figsize=(11, 6))
    for name, (tt, ll, correction) in traces.items():
        if name in selected_available:
            ax.plot(tt, ll, lw=1.0, alpha=0.75,
                    label=f"{DISPLAY_NAMES[name]} (+{correction:.2f} dB)")
        elif SOURCE_ESTIMATE_SHOW_EXCLUDED_CHANNELS:
            ax.plot(tt, ll, lw=0.8, ls="--", alpha=0.35,
                    label=f"QC-excluded: {DISPLAY_NAMES[name]}")

    if len(consensus_t):
        ax.fill_between(consensus_t, p10, p90, alpha=0.16,
                        label="10–90% spread of selected channels")
        ax.plot(consensus_t, consensus, lw=2.2,
                label="QC-selected median source-equivalent estimate")

    ax.axvline(0.0, ls=":", alpha=0.7, label="Detected acoustic onset")
    ax.set_xlabel("Event-relative time (s)")
    ax.set_ylabel(f"Equivalent SPL at {SOURCE_EQUIVALENT_REFERENCE_M:g} m (dB re 20 µPa)")
    ax.set_title(
        f"19. Rough Source-Equivalent SPL vs Time — {SOURCE_EQUIVALENT_REFERENCE_M:g} m Reference\n"
        "Spherical free-field back-propagation; not literal SPL at r = 0"
    )
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=8, ncol=2)
    savefig(fig, outdir/"19_estimated_source_equivalent_SPL_1m_vs_time.png")

    # Individual source-equivalent companion graphs.
    individual_dir = get_individual_plot_dir(outdir)
    for name, (tt, ll, correction) in traces.items():
        fig_ind, ax_ind = plt.subplots(figsize=(11, 6))
        ax_ind.plot(tt, ll, lw=1.1)
        ax_ind.axvline(0.0, ls=":", alpha=0.7, label="Detected acoustic onset")
        ax_ind.set_xlabel("Event-relative time (s)")
        ax_ind.set_ylabel(
            f"Equivalent SPL at {SOURCE_EQUIVALENT_REFERENCE_M:g} m (dB re 20 µPa)"
        )
        ax_ind.set_title(
            f"19. Source-Equivalent SPL — {DISPLAY_NAMES[name]}\n"
            f"Measured at {DISTANCES_M[name]:g} m, corrected by +{correction:.2f} dB"
        )
        ax_ind.grid(True, alpha=0.3)
        ax_ind.legend()
        savefig(
            fig_ind,
            individual_dir/f"19_source_equivalent_1m_{name}.png"
        )

    detail_df = pd.DataFrame(rows)
    detail_df.to_csv(tables/"estimated_source_equivalent_SPL_1m_all_channels.csv", index=False)

    consensus_df = pd.DataFrame({
        "event_relative_time_s": consensus_t,
        "source_equivalent_SPL_1m_median_db": consensus,
        "selected_channel_10th_percentile_db": p10,
        "selected_channel_90th_percentile_db": p90,
    })
    consensus_df.to_csv(tables/"estimated_source_equivalent_SPL_1m_consensus.csv", index=False)

    selection_rows = [{
        "selected_channels": ", ".join(selected_available),
        "selection_note": selection_note,
        "reference_distance_m": SOURCE_EQUIVALENT_REFERENCE_M,
        "model": "L_ref = L_r + 20 log10(r/ref); spherical free-field approximation",
        "warning": "Not literal r=0 source SPL; ignores directivity, ground reflection and distributed plume source",
        "rejected_pair_details": " | ".join(f"{k}: {v}" for k, v in rejection_reasons.items()),
    }]
    pd.DataFrame(selection_rows).to_csv(tables/"source_estimate_qc.csv", index=False)
    return consensus_df, selected_available, selection_note


def paper_style_plots(results, spectra, third, metrics, outdir):
    individual_dir = get_individual_plot_dir(outdir)

    # Smoothed narrowband SPL (paper-style): combined + individual.
    fig, ax = plt.subplots(figsize=(10, 6))
    for name, (f, P) in spectra.items():
        df = np.median(np.diff(f))
        L = db10(P*df/P_REF**2)

        win = min(SAVGOL_WINDOW_POINTS, len(L) if len(L)%2 else len(L)-1)
        if win >= 5 and win > SAVGOL_POLYORDER:
            if win % 2 == 0:
                win -= 1
            L = signal.savgol_filter(L, win, SAVGOL_POLYORDER)

        ax.semilogx(f[1:], L[1:], label=DISPLAY_NAMES[name])

        fig_ind, ax_ind = plt.subplots(figsize=(10, 6))
        ax_ind.semilogx(f[1:], L[1:], lw=1.0)
        ax_ind.set_xlim(FREQ_MIN_HZ, FREQ_MAX_HZ)
        ax_ind.set_xlabel("Frequency (Hz)")
        ax_ind.set_ylabel("Narrowband SPL per Welch bin (dB re 20 µPa)")
        ax_ind.set_title(
            f"15. Paper-Style Smoothed Narrowband SPL — {DISPLAY_NAMES[name]}"
        )
        ax_ind.grid(True, which="both", alpha=0.3)
        savefig(fig_ind, individual_dir/f"15_paper_narrowband_{name}.png")

    ax.set_xlim(FREQ_MIN_HZ, FREQ_MAX_HZ)
    ax.set_xlabel("Frequency (Hz)")
    ax.set_ylabel("Narrowband SPL per Welch bin (dB re 20 µPa)")
    ax.set_title("15. Paper-Style Smoothed Narrowband SPL — All Sensors")
    ax.grid(True, which="both", alpha=0.3)
    ax.legend()
    savefig(fig, outdir/"15_paper_style_smoothed_narrowband_spl.png")

    # Individual unweighted and A-weighted 1/3 octave.
    for name, df13 in third.items():
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.semilogx(df13["center_hz"], df13["spl_db"], "o-", label="Unweighted")
        ax.semilogx(
            df13["center_hz"], df13["A_weighted_spl_dbA"],
            "s-", label="A-weighted"
        )
        ax.set_xlabel("1/3-octave centre frequency (Hz)")
        ax.set_ylabel("Band SPL (dB)")
        ax.set_title(f"16. 1/3-Octave Unweighted / A-Weighted — {DISPLAY_NAMES[name]}")
        ax.grid(True, which="both", alpha=0.3)
        ax.legend()
        savefig(fig, outdir/f"16_third_octave_Aweighted_{name}.png")

        # Fasulo-style reflection correction range.
        lo, hi = REFLECTION_CORRECTION_DB
        y = df13["spl_db"].to_numpy()
        x = df13["center_hz"].to_numpy()

        fig, ax = plt.subplots(figsize=(10, 6))
        ax.semilogx(x, y, label="Measured")
        ax.fill_between(
            x, y-hi, y-lo, alpha=0.3,
            label=f"Illustrative −{lo:g} to −{hi:g} dB reflection sensitivity band"
        )
        ax.set_xlabel("1/3-octave centre frequency (Hz)")
        ax.set_ylabel("Band SPL (dB re 20 µPa)")
        ax.set_title(f"17. Illustrative Reflection Sensitivity Band — {DISPLAY_NAMES[name]}\n(not a measured site-specific correction)")
        ax.grid(True, which="both", alpha=0.3)
        ax.legend()
        savefig(fig, outdir/f"17_reflection_correction_{name}.png")

    # Fitted spreading coefficient: L = A - k log10(r)
    m = metrics.sort_values("distance_m")
    x = np.log10(m["distance_m"].to_numpy())
    y = m["rms_spl_db"].to_numpy()
    slope, intercept = np.polyfit(x, y, 1)
    k = -slope
    pred = slope*x + intercept
    ss_res = np.sum((y-pred)**2)
    ss_tot = np.sum((y-np.mean(y))**2)
    r2 = 1 - ss_res/ss_tot if ss_tot > 0 else np.nan

    pd.DataFrame([{
        "equivalent_1m_rms_level_db": intercept,
        "measured_spreading_coefficient_k": k,
        "ideal_spherical_coefficient_k": 20.0,
        "R_squared": r2,
    }]).to_csv(outdir/"tables"/"measured_spreading_fit.csv", index=False)


# ============================================================================
# OUTPUT-FOLDER MANAGEMENT
# ============================================================================

def create_new_run_folder(script_folder: Path) -> Path:
    """
    Create a brand-new sequential results folder for every execution.

    Examples:
        acoustic_results_run_001
        acoustic_results_run_002
        acoustic_results_run_003

    Existing folders are NEVER reused or overwritten.
    """
    pattern = re.compile(
        rf"^{re.escape(OUTPUT_FOLDER_NAME)}_(\d+)$",
        flags=re.IGNORECASE
    )

    existing_numbers = []
    for p in script_folder.iterdir():
        if not p.is_dir():
            continue
        m = pattern.match(p.name)
        if m:
            try:
                existing_numbers.append(int(m.group(1)))
            except ValueError:
                pass

    run_number = max(existing_numbers, default=0) + 1

    # Extremely defensive: keep incrementing if a path somehow already exists.
    while True:
        outdir = script_folder / f"{OUTPUT_FOLDER_NAME}_{run_number:03d}"
        if not outdir.exists():
            outdir.mkdir(parents=True, exist_ok=False)
            return outdir
        run_number += 1


def write_plot_index(outdir: Path):
    """Create a simple text index so every output is easy to identify."""
    content = """HYBRID ROCKET ACOUSTIC ANALYSIS — PLOT INDEX
================================================

01  Calibrated acoustic pressure vs time — all sensors
02  SPL vs time — all sensors
03  Robust peak SPL vs distance
04  RMS SPL vs distance
05  Measured attenuation vs theoretical 20 log10(r2/r1)
06  PSD — all sensors
07  Near-array vs far-array average PSD
08  Spectral attenuation: near array -> far array
09  1/3-octave SPL — near vs far
10  Spectrogram — one graph for each sensor
11  Cross-correlation — near 1 m pair
12  Cross-correlation — far 1 m pair
13  Coherence vs frequency — both 1 m pairs
14  Distance-corrected spectra
15  Paper-style smoothed narrowband spectrum
16  1/3-octave unweighted + A-weighted — one graph per sensor
17  Reflection-correction range — one graph per sensor
18  SPL vs chamber pressure — one graph per sensor
18a Chamber-pressure alignment reference
19  Estimated source-equivalent SPL at 1 m vs time
20  Direct single-sided FFT amplitude — all sensors

INDIVIDUAL COMPANION PLOTS
--------------------------
The folder:
    individual_sensor_plots/

contains separate per-sensor versions of the combined sensor graphs, including:

01  Acoustic pressure vs time — each sensor separately
02  SPL vs time — each sensor separately
06  PSD — each sensor separately
14  Distance-corrected spectrum — each sensor separately
15  Paper-style narrowband spectrum — each sensor separately
19  Source-equivalent SPL at 1 m — each sensor separately
20  FFT — each sensor separately

Plots that are inherently pair/group/distance comparisons (03, 04, 05, 07, 08,
09, 11, 12, 13) remain as comparison graphs. Plots 10, 16, 17 and 18 were already
generated individually per sensor.

Additional CSV tables are in:
    tables/

IMPORTANT:
- Plot 19 is an equivalent free-field level back-propagated to 1 m.
  It is NOT literal pressure at r = 0.
- For broadband rocket noise, PSD (plot 06) is generally better for comparing
  spectral energy; FFT (plot 20) is especially useful for identifying tones,
  resonances and dominant frequencies.
"""
    (outdir / "00_PLOT_INDEX.txt").write_text(content, encoding="utf-8")


# ============================================================================
# MAIN
# ============================================================================

def main():
    script_folder = Path(__file__).resolve().parent

    # IMPORTANT: create a completely new folder for every run.
    outdir = create_new_run_folder(script_folder)
    tables = outdir / "tables"
    tables.mkdir(parents=True, exist_ok=False)
    write_plot_index(outdir)

    print(f"NEW results folder:   {outdir.name}")
    print("Existing result folders will not be overwritten.")

    print("="*78)
    print("HYBRID ROCKET ACOUSTIC ANALYSIS — V7 COMBINED + INDIVIDUAL PLOTS / UNIQUE RUN FOLDERS")
    print("="*78)
    print(f"Sampling rate:       {SAMPLE_RATE_HZ:,.0f} Hz")
    print(f"Maximum samples:     {MAX_SAMPLES:,} per channel")
    print(f"Maximum duration:    {MAX_SAMPLES/SAMPLE_RATE_HZ:.3f} s")
    print(f"Analysis band:       {FILTER_LOW_HZ:g}–{FILTER_HIGH_HZ:g} Hz")
    print()

    test1, test2 = discover_input_files(script_folder)
    print(f"Test 1 file: {test1.name}")
    print(f"Test 2 file: {test2.name}")

    # Read acoustic data.
    c11, c12, t1raw, s11, s12 = read_waveform(test1, MAX_SAMPLES)
    c21, c22, t2raw, s21, s22 = read_waveform(test2, MAX_SAMPLES)

    n1 = min(len(c11), len(c12), MAX_SAMPLES)
    n2 = min(len(c21), len(c22), MAX_SAMPLES)
    c11, c12 = c11[:n1], c12[:n1]
    c21, c22 = c21[:n2], c22[:n2]

    # Fail-fast QC catches Time-as-channel and other obvious acquisition mistakes.
    qc_raw = {
        "T1_CH1": validate_raw_acoustic_signal(c11, "Test 1 CH1"),
        "T1_CH2": validate_raw_acoustic_signal(c12, "Test 1 CH2"),
        "T2_CH1": validate_raw_acoustic_signal(c21, "Test 2 CH1"),
        "T2_CH2": validate_raw_acoustic_signal(c22, "Test 2 CH2"),
    }

    t1 = time_axis(t1raw, n1)
    t2 = time_axis(t2raw, n2)

    print(f"\nTest 1 simultaneous samples: {n1:,} ({n1/SAMPLE_RATE_HZ:.4f} s)")
    print(f"Test 2 simultaneous samples: {n2:,} ({n2/SAMPLE_RATE_HZ:.4f} s)")

    raw_signals = {
        "T1_CH1": (t1, c11, s11),
        "T1_CH2": (t1, c12, s12),
        "T2_CH1": (t2, c21, s21),
        "T2_CH2": (t2, c22, s22),
    }

    # Detect hard acquisition dropouts BEFORE detrending/filtering. A zero-filled
    # tail must never be treated as genuine quiet acoustic data.
    signals = {}
    qc_rows = []
    for name, (t, raw, vscale) in raw_signals.items():
        keep_n, flat_info = detect_hard_flatline_tail(raw)
        raw_valid = raw[:keep_n]
        t_valid = t[:keep_n]
        if keep_n < len(raw):
            print(
                f"WARNING: {name} hard flatline/dropout detected at "
                f"{keep_n/SAMPLE_RATE_HZ:.3f} s; trimming "
                f"{len(raw)-keep_n:,} invalid tail samples."
            )

        p_cal = calibrate_to_pa(raw_valid, name, vscale)
        p_clean, clean_info = preprocess_acoustic_pressure(p_cal)
        signals[name] = (t_valid, p_clean)

        abs_p = np.abs(p_clean)
        robust = float(np.percentile(abs_p, ROBUST_PEAK_PERCENTILE)) if len(abs_p) else np.nan
        absolute = float(np.max(abs_p)) if len(abs_p) else np.nan
        spike_ratio = absolute/max(robust, EPS) if np.isfinite(robust) else np.nan
        status_bits = []
        if flat_info["flatline_tail_detected"]:
            status_bits.append("PARTIAL RECORD — HARD DROPOUT")
        if np.isfinite(spike_ratio) and spike_ratio > 2.0:
            status_bits.append("CHECK IMPULSIVE SPIKES")
        if not status_bits:
            status_bits.append("OK")

        row = {
            "sensor": name,
            "description": DISPLAY_NAMES[name],
            **qc_raw[name],
            **flat_info,
            **clean_info,
            "original_samples": len(raw),
            "valid_samples_after_trim": keep_n,
            "valid_duration_s": keep_n/SAMPLE_RATE_HZ,
            "coverage_fraction": keep_n/max(len(raw), 1),
            "filtered_std_pa": float(np.std(p_clean)),
            "absolute_to_robust_peak_ratio": spike_ratio,
            "channel_status": "; ".join(status_bits),
        }
        qc_rows.append(row)

    qc_df = pd.DataFrame(qc_rows)
    qc_df.to_csv(tables/"signal_quality_control.csv", index=False)
    print("\nSignal QC:")
    print(qc_df[[
        "sensor", "valid_duration_s", "flatline_tail_detected",
        "absolute_to_robust_peak_ratio", "channel_status"
    ]].to_string(index=False))

    # Full valid traces are used for time histories / spectrograms / chamber overlay.
    full_results = {}
    for name, (t, p) in signals.items():
        st, sdb = sliding_spl(t, p)
        full_results[name] = {"t": t, "p": p, "spl_t": st, "spl_db": sdb}

    # Cross-distance comparisons MUST use equal event-relative windows.
    comparison_segments, comparison_window_df, acoustic_onsets = build_common_event_segments(signals)
    comparison_window_df.to_csv(tables/"common_comparison_window.csv", index=False)

    comparison_results = {}
    spectra = {}
    third = {}
    metric_rows = []
    for name, p in comparison_segments.items():
        tcmp = np.arange(len(p), dtype=float)/SAMPLE_RATE_HZ
        st, sdb = sliding_spl(tcmp, p)
        comparison_results[name] = {"t": tcmp, "p": p, "spl_t": st, "spl_db": sdb}

        f, P = welch_psd(p)
        spectra[name] = (f, P)
        third[name] = third_octave_from_psd(f, P)
        third[name].to_csv(tables/f"third_octave_{name}.csv", index=False)

        p_rms = float(np.sqrt(np.mean(p*p)))
        p_peak = float(np.max(np.abs(p)))
        robust_peak = float(np.percentile(np.abs(p), ROBUST_PEAK_PERCENTILE))

        metric_rows.append({
            "sensor": name,
            "description": DISPLAY_NAMES[name],
            "distance_m": DISTANCES_M[name],
            "comparison_samples": len(p),
            "comparison_duration_s": len(p)/SAMPLE_RATE_HZ,
            "sample_rate_hz": SAMPLE_RATE_HZ,
            "absolute_max_pressure_pa_diagnostic": p_peak,
            "robust_peak_pressure_pa": robust_peak,
            "rms_pressure_pa": p_rms,
            "absolute_max_spl_db_diagnostic": float(safe_pressure_to_spl(p_peak)),
            "robust_peak_spl_db": float(safe_pressure_to_spl(robust_peak)),
            "rms_spl_db": float(safe_pressure_to_spl(p_rms)),
            "OASPL_20Hz_to_20kHz_db": integrate_psd_level(f, P, weighted=False),
            "A_weighted_OASPL_20Hz_to_20kHz_dbA": integrate_psd_level(f, P, weighted=True),
        })

    metrics = pd.DataFrame(metric_rows).sort_values("distance_m")
    metrics.to_csv(tables/"summary_metrics_common_window.csv", index=False)
    # Keep the conventional filename too so existing workflow still finds it.
    metrics.to_csv(tables/"summary_metrics.csv", index=False)

    # Requested plots 1–6, now based on the cleaned 20 Hz–20 kHz acoustic waveform.
    plot_pressure_time(full_results, outdir)
    plot_spl_time(full_results, outdir)
    plot_distance_metric(
        metrics, "robust_peak_spl_db",
        f"3. Robust Peak SPL vs Distance ({ROBUST_PEAK_PERCENTILE:.2f}th percentile; common window)",
        "03_peak_spl_vs_distance.png", outdir
    )
    plot_distance_metric(
        metrics, "rms_spl_db",
        "4. RMS SPL vs Distance (20 Hz–20 kHz; common ignition-aligned window)",
        "04_rms_spl_vs_distance.png", outdir
    )
    plot_attenuation(metrics, outdir)
    plot_all_psd(comparison_results, spectra, outdir)

    # Requested direct FFT. This uses the same common ignition-aligned comparison
    # window as PSD/attenuation so every sensor has equal record length and bin width.
    fft_df = plot_fft_all_sensors(comparison_results, outdir, tables)

    # Attenuation table.
    rows = []
    m = metrics.sort_values("distance_m").reset_index(drop=True)
    for i in range(len(m)):
        for j in range(i+1, len(m)):
            r1, r2 = m.loc[i, "distance_m"], m.loc[j, "distance_m"]
            L1, L2 = m.loc[i, "rms_spl_db"], m.loc[j, "rms_spl_db"]
            theory = 20*np.log10(r2/r1)
            measured = L1-L2
            rows.append({
                "sensor_1": m.loc[i, "sensor"],
                "sensor_2": m.loc[j, "sensor"],
                "r1_m": r1,
                "r2_m": r2,
                "measured_rms_attenuation_db": measured,
                "theoretical_20log10_r2_r1_db": theory,
                "measured_minus_theoretical_db": measured-theory,
            })
    pd.DataFrame(rows).to_csv(tables/"attenuation_all_pairs.csv", index=False)

    # Distance-comparison QC: if the far array is louder than the near array in the
    # matched window, do not interpret the result as physical propagation loss.
    near_levels = metrics[metrics["sensor"].isin(NEAR_SENSORS)]["rms_spl_db"].to_numpy(float)
    far_levels = metrics[metrics["sensor"].isin(FAR_SENSORS)]["rms_spl_db"].to_numpy(float)
    near_group_db = float(10*np.log10(np.mean(10**(near_levels/10.0))))
    far_group_db = float(10*np.log10(np.mean(10**(far_levels/10.0))))
    measured_group_attenuation = near_group_db - far_group_db
    expected_group_attenuation = 20*np.log10(
        np.mean([DISTANCES_M[x] for x in FAR_SENSORS]) /
        np.mean([DISTANCES_M[x] for x in NEAR_SENSORS])
    )
    distance_status = "OK FOR ATTENUATION INTERPRETATION"
    if measured_group_attenuation < 0:
        distance_status = (
            "FAIL QC — FAR ARRAY LOUDER THAN NEAR ARRAY; CHECK CALIBRATION, "
            "GAIN, SENSOR HEALTH, AND WHETHER DATASETS REPRESENT THE SAME SOURCE EVENT"
        )
        warnings.warn(distance_status)
    pd.DataFrame([{
        "near_group_energy_average_rms_spl_db": near_group_db,
        "far_group_energy_average_rms_spl_db": far_group_db,
        "measured_near_minus_far_db": measured_group_attenuation,
        "expected_spherical_near_minus_far_db": expected_group_attenuation,
        "distance_comparison_status": distance_status,
    }]).to_csv(tables/"distance_comparison_qc.csv", index=False)

    # Requested 7–9.
    fgrp, Pnear, Pfar, rnear, rfar = plot_near_far_psd(spectra, outdir)
    plot_spectral_attenuation(fgrp, Pnear, Pfar, rnear, rfar, outdir)
    plot_third_octave_near_far(
        fgrp, Pnear, Pfar, rnear, rfar, outdir, tables
    )

    # Requested 10–14.
    plot_spectrograms(full_results, outdir)
    pair_df = plot_pairs(comparison_results, outdir, tables)
    plot_distance_corrected(spectra, outdir)

    # Consolidated sensor-health verdict. This does not delete questionable data;
    # it clearly flags channels/pairs so bad hardware/acquisition is not mistaken
    # for rocket acoustics.
    health_df = build_sensor_health_summary(qc_df, pair_df, metrics)
    health_df.to_csv(tables/"sensor_health_summary.csv", index=False)
    print("\nSensor health summary:")
    print(health_df[["sensor", "health", "notes"]].to_string(index=False))

    # New: rough 1 m source-equivalent SPL estimate.  The script selects the
    # healthiest simultaneous pair using dropout/coherence/delay QC.
    source_df, source_channels, source_note = plot_source_equivalent_spl(
        full_results, qc_df, pair_df, outdir, tables
    )

    # Paper-inspired outputs.
    paper_style_plots(comparison_results, spectra, third, metrics, outdir)

    # New: chamber-pressure / SPL time alignment.
    chamber = load_chamber_pressure(script_folder)
    plot_chamber_pressure_alignment(full_results, chamber, outdir, tables)

    # Human-readable summary.
    with (outdir/"READ_ME_RESULTS.txt").open("w", encoding="utf-8") as f:
        f.write("HYBRID ROCKET ACOUSTIC RESULTS — V7 COMBINED + INDIVIDUAL PLOTS / UNIQUE RUN FOLDERS\n")
        f.write("="*52 + "\n\n")
        f.write(f"RESULTS FOLDER: {outdir.name}\\n")
        f.write("This folder was created uniquely for this run; older calculations were not overwritten.\\n\\n")

        f.write("IMPORTANT FIX\\n")
        f.write(
            "The previous CSV parser could analyse the Time column as Channel 1 "
            "because pandas returned usecols in file order. This version selects "
            "Time, Channel 1 and Channel 2 explicitly by column name and stops if "
            "a channel looks like a monotonic time/ramp signal.\n\n"
        )
        f.write(f"Sampling rate: {SAMPLE_RATE_HZ:,.0f} Hz\n")
        f.write(f"Configured max samples: {MAX_SAMPLES:,}\n")
        f.write(f"Configured max duration: {MAX_SAMPLES/SAMPLE_RATE_HZ:.3f} s\n")
        f.write(f"Analysis band: {FILTER_LOW_HZ:g}–{FILTER_HIGH_HZ:g} Hz\n\n")
        f.write("Sensor mapping:\n")
        for k in DISTANCES_M:
            f.write(f"  {k}: {DISPLAY_NAMES[k]}\n")
        f.write("\n")
        f.write(metrics.to_string(index=False))
        f.write("\n\nSIGNAL QUALITY CONTROL\n")
        f.write(qc_df[["sensor", "valid_duration_s", "flatline_tail_detected", "absolute_to_robust_peak_ratio", "channel_status"]].to_string(index=False))
        f.write("\n\nPAIR QUALITY CONTROL\n")
        f.write(pair_df.to_string(index=False))
        f.write("\n\nDISTANCE COMPARISON QC\n")
        f.write(
            f"Near energy-average RMS SPL: {near_group_db:.2f} dB\n"
            f"Far energy-average RMS SPL: {far_group_db:.2f} dB\n"
            f"Measured near-minus-far: {measured_group_attenuation:.2f} dB\n"
            f"Expected spherical attenuation: {expected_group_attenuation:.2f} dB\n"
            f"Status: {distance_status}\n"
        )
        f.write("\n")
        f.write("Static-fire chamber pressure:\n")
        if chamber:
            f.write(
                f"  File: {CHAMBER_PRESSURE_FILE}\n"
                f"  Column: {CHAMBER_PRESSURE_COLUMN}\n"
                f"  Detected DAQ ignition time: "
                f"{chamber['onset_time_daq_s']:.3f} s\n"
                f"  Pressure unit: {CHAMBER_PRESSURE_UNIT} "
                "(workbook did not specify bar/psi)\n"
            )
            f.write(
                "  Acoustic and pressure clocks are not shared. Each acoustic "
                "channel is therefore aligned to the pressure trace by matching "
                "the ignition event, not by trusting absolute timestamps.\n"
            )
        else:
            f.write("  Chamber-pressure file not available; alignment skipped.\n")

        f.write("\nSOURCE-EQUIVALENT ESTIMATE\n")
        f.write(
            f"Reference distance: {SOURCE_EQUIVALENT_REFERENCE_M:g} m\n"
            f"Channels used for consensus: {', '.join(source_channels)}\n"
            f"Selection: {source_note}\n"
            "Method: L_ref = L_measured + 20 log10(r/ref). This is only a rough "
            "spherical free-field back-propagation and is NOT literal SPL at r=0.\n"
        )

        f.write("\nFFT output:\n")
        f.write("  Plot 20: 20_fft_amplitude_all_sensors.png\n")
        f.write("  Table: tables/fft_amplitude_all_sensors.csv\n")
        f.write("  Note: for broadband rocket noise, PSD is the preferred level comparison because FFT amplitude depends on FFT-bin width.\n")

        f.write("\nCalibration note:\n")
        f.write(
            "Absolute SPL is only as accurate as the per-channel microphone "
            "sensitivity and conditioner gain in CHANNEL_CALIBRATION.\n"
        )

    print("\nAnalysis complete.")
    print(f"Results folder: {outdir}")
    print(f"Summary table:  {tables/'summary_metrics.csv'}")
    if chamber:
        print(
            "Pressure alignment: generated one SPL/chamber-pressure overlay "
            "for each of the four acoustic channels."
        )
    print(
        "\nGenerated the COMPLETE graph set: plots 1–18 from the previous workflow, "
        "plus plot 19 source-equivalent SPL and plot 20 direct FFT."
    )
    print("Combined plots are in the main run folder.")
    print(f"Individual sensor companion plots are in: {outdir/'individual_sensor_plots'}")
    print(f"Nothing from previous runs was overwritten. New run folder: {outdir.name}")
    print(f"See {outdir/'00_PLOT_INDEX.txt'} for the full graph list.")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("\nERROR")
        print("-----")
        print(str(exc))
        print(
            "\nIf this says no waveform table was detected, the file is probably "
            "a WaveForms Measurements export rather than the raw Scope data export."
        )
        raise
